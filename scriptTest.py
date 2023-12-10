import requests
import os
import re
import openpyxl
import unidecode

from bs4 import BeautifulSoup
from urllib.parse import urljoin

"""
1) Identificar correctamente el elemento(s) que contienen 
informacion deseada
   Elemento general: 
    <div id="article_i__ss_ar_articulo_archivos_1" class="recuadros download-attr margen-abajo-md"><div class="recuadro format-pdf externo cid-507 aid-708568 binary-archivo_01 format-xlsx media"><a href="articles-708568_archivo_01.xlsx" title="Ir a Estadísticas de la Seguridad Social 2022" download="Estadísticas de la Seguridad Social 2022.pdf">Estadísticas de la Seguridad Social 2022</a></div></div>
 Especificamete:
    div que contiene el link de descarga > article_i__ss_ar_articulo_archivos_1
    <a href="articles-708568_archivo_01.xlsx" title="Ir a Estadísticas de la Seguridad Social 2022" download="Estadísticas de la Seguridad Social 2022.pdf">Estadísticas de la Seguridad Social 2022</a>

"""
def load_downloaded_files(recordFile):
    try:
        with open(recordFile, 'r', encoding='utf-8') as file:
            downloadedFiles = [tuple(line.strip().split(',')) for line in file.readlines()]
        return downloadedFiles
    except FileNotFoundError:
        return []

def save_downloaded_files(recordFile, downloadedFiles):
    with open(recordFile, 'w', encoding='utf-8') as file:
        for title, filePath in downloadedFiles:
            file.write(f"{title},{filePath}\n")



def downloadExcelFiles(url, downloadDir, recordFile='downloadedFiles.txt'):
    resp = requests.get(url)
    soup = BeautifulSoup(resp.text, 'html.parser')
    
    os.makedirs(downloadDir, exist_ok=True)
    
    # Find the table with id 'tabular_generica'
    table = soup.find('table', {'id': 'tabular_generica'})
    
    downloadedFiles = load_downloaded_files(recordFile)
    
 
    if table:
        # Find all rows in the table
        rows = table.find_all('tr')
        for row in rows:
            # Find all cells in the row
            cells = row.find_all(['td', 'th'])
            for cell in cells:
                # Check if the cell contains a link
                link = cell.find('a', href=True)
                if link:
                    downloadUrl = link['href']
                    title = link.get('title', '')
                    if downloadUrl.endswith('.xlsx') and 'Seguridad Social' in title:
                        fullUrl = urljoin(url, downloadUrl)
                        
                        #limpieza 
                        newName = f"{title.replace(' ', '_')}download.xlsx"
                        filePath = os.path.join(downloadDir, newName)
                
                        
                        if not os.path.isfile(filePath):
                            with open(filePath, 'wb') as file:
                                file.write(requests.get(fullUrl).content)
                            
                            downloadedFiles.append((title, filePath))
                            print(f"Download: {newName}")
                        else:
                            print(f"Skipped download for {newName}. File already exists / El archivo ya existe.")
    save_downloaded_files(recordFile, downloadedFiles)
    return list(downloadedFiles)     
    
def extractedDataFromExcel_Type1(filePath):  
    wb = openpyxl.load_workbook(filePath)
    sheetsToProcess = ['31', '29', '38', '39', '28']
    for wantedSheet in sheetsToProcess:
        if  wantedSheet in wb.sheetnames:
            ws = wb[wantedSheet]
            if wantedSheet == '31':       
                print(f"switched to sheet: {ws.title}")             
                categoryData = {
                        "ACCIDENTES DEL TRABAJO": {
                            "economicActivityStart": 'B9',
                            "economicActivityEnd": 'B26',
                            "totalColumn": 'F',
                        },
                        "ACCIDENTES DE TRAYECTO": {
                            "economicActivityStart": 'B28',
                            "economicActivityEnd": 'B45',  # Updated to 'B45'
                            "totalColumn": 'F',
                        },
                        "ACCIDENTES (TRABAJO + TRAYECTO)": {
                            "economicActivityStart": 'B48',  # Updated to 'B48'
                            "economicActivityEnd": 'B64',
                            "totalColumn": 'F',
                        },
                }

                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    total_column = categoryInfo['totalColumn']
                    
                    economicActivities = []
                    total_values = []

                    for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                        economic_activity_cell = ws.cell(row=row, column=economicActivityStart.column)
                        economic_activity = economic_activity_cell.value
                        
                        achs_value = economic_activity_cell.offset(column=1).value
                        museg_value = economic_activity_cell.offset(column=2).value
                        ist_value = economic_activity_cell.offset(column=3).value
                        total_value = ws[f"{total_column}{row}"].value  # Fetch the "Total" from the specified column
                        
                        economicActivities.append({
                            "Economic Activity": economic_activity,
                            "ACHS": achs_value,
                            "MUSEG": museg_value,
                            "IST": ist_value,
                        })
                        total_values.append(total_value)

                    data[category]['Total'] = total_values  # Store all "Total" values for each economic activity
                    data[category]['Economic Activities'] = economicActivities

                for category, categoryInfo in data.items():
                    print(category)
                    for i, economic_activity in enumerate(categoryInfo['Economic Activities']):
                        print(f"Economic activity: {economic_activity['Economic Activity']}")
                        print(f"ACHS: {economic_activity['ACHS']}")
                        print(f"MUSEG: {economic_activity['MUSEG']}")
                        print(f"IST: {economic_activity['IST']}")
                        print(f"Total: {categoryInfo['Total'][i]}")
                    print()
            if wantedSheet == '29':
                print(f"switched to sheet: {ws.title}")
                categoryData = {
                    "ACCIDENTES DEL TRABAJO": {
                        "mutualInicio": "B8",
                        "mutualFinal": "B11",

                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "mutualInicio": "B13",
                        "mutualFinal": "B16",
                        
                    },
                    "POR ACCIDENTES (TRABAJO + TRAYECTO)": {
                        "mutualInicio": "B18",
                        "mutualFinal": "B21",
                    
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    mutualInicioCell = ws[categoryInfo["mutualInicio"]]
                    mutualFinalCell = ws[categoryInfo["mutualFinal"]]
                    
        
                    
                    mutuales = []

                    for row in range(mutualInicioCell.row, mutualFinalCell.row + 1):
                        mutualInicioCell = ws.cell(row=row, column=mutualInicioCell.column)
                        mutualInicio = mutualInicioCell.value
                        
                        anio2018 = mutualInicioCell.offset(column=1).value
                        anio2019 = mutualInicioCell.offset(column=2).value
                        anio2020 = mutualInicioCell.offset(column=3).value
                        anio2021 = mutualInicioCell.offset(column=4).value
                        anio2022 = mutualInicioCell.offset(column=5).value

                        mutuales.append({
                            "Mutualidades": mutualInicio,
                            "2018": anio2018,
                            "2019": anio2019,
                            "2020": anio2020,
                            "2021": anio2021,
                            "2022": anio2022
                        })
                    data[category]['Mutuales'] = mutuales
                        
                for category, categoryInfo in data.items():
                    print(category)
                    for i, mutualInicio in enumerate(categoryInfo['Mutuales']):
                        print(f"mutuales: {mutualInicio['Mutualidades']}")
                        print(f"2018: {mutualInicio['2018']}")
                        print(f"2019: {mutualInicio['2019']}")
                        print(f"2020: {mutualInicio['2020']}")
                        print(f"2021: {mutualInicio['2021']}")
                        print(f"2022: {mutualInicio['2022']}")
                    print()
            if wantedSheet == '38':
                print(f"switched to sheet: {ws.title}")
                
                categoryData = {
                    "ACCIDENTES DEL TRABAJO":{
                        "economicActivityStart": 'B8',
                        "economicActivityEnd": 'B24',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "economicActivityStart": 'B26',
                        "economicActivityEnd":'B42',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)":{
                        "economicActivityStart":'B44',
                        "economicActivityEnd": 'B60',
                        "totalColumn": 'G'
                    }       
                }
                    
                data = {}
                
                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']  
                    
                    economicActivities = []
                    totalValues = []
                    
                    for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
                        economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value
                        
                        achsValue = economicActivityCell.offset(column=1).value
                        musegValue = economicActivityCell.offset(column=2).value
                        istValue = economicActivityCell.offset(column=3).value
                        islValue = economicActivityCell.offset(column=4).value
                        totalValue = ws[f"{totalColumn}{row}"].value
                        
                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "ACHS": achsValue,
                            "MUSEG": musegValue,
                            "IST": istValue,
                            "ISL": islValue
                        })
                        totalValues.append(totalValue)
                        
                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities
                for category, categoryInfo in data.items():
                    print(category)
                    for i, economicActivity in enumerate(categoryInfo['Economic Activities']):
                        print(f"Economic activity:{economicActivity['Economic Activity']}")
                        print(f"ACHS: {economicActivity['ACHS']}")
                        print(f"MUSEG: {economicActivity['MUSEG']}")
                        print(f"IST: {economicActivity['IST']}")
                        print(f"ISL: {economicActivity['ISL']}")
                        print(f"Total: {categoryInfo['Total'][i]}")
                    print()

            if wantedSheet == '39':
                print(f"switched to sheet: {ws.title}")
                categoryData = {
                    "ACCIDENTES DEL TRABAJO": {
                        "economicActivityStart": 'B8',
                        "economicActivityEnd": 'B24',
                        "totalColumn": 'E'
                    },
                    "ACCIDENTES DEL TRAYECTO": {
                        "economicActivityStart": 'B26',
                        "economicActivityEnd": 'B42',
                        "totalColumn": 'E'
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)":{
                        "economicActivityStart": 'B44',
                        "economicActivityEnd": 'B60',
                        "totalColumn": 'E'
                    }
                }
                
                data = {}
                
                for category, categoryInfo in categoryData.items():
                    
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']
                    economicActivities = []
                    totalValues = []
                    
                    for row in range (economicActivityStart.row, economicActivityEnd.row+1):
                        economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value
                        
                        menValues = economicActivityCell.offset(column=1).value
                        womenValue = economicActivityCell.offset(column=2).value
                        totalValue = ws[f"{totalColumn}{row}"].value
                        
                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "Men": menValues,
                            "Women": womenValue
                        })
                        totalValues.append(totalValue)
                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities
                for category, categoryInfo in data.items():
                    print(category)
                    for i, economicActivity in enumerate(categoryInfo['Economic Activities']):
                        print(f"Economic Activity: {economicActivity['Economic Activity']}")
                        print(f"Men: {economicActivity['Men']}")
                        print(f"Women: {economicActivity['Women']}")
                        print(f"Total: {categoryInfo['Total'][i]}")
            if wantedSheet == '28':
                print(f"switched to sheet: {ws.title}")
                categoryData = {
                    "ACCIDENTES DEL TRABAJO": {
                        "economicActivityStart": 'B9',
                        "economicActivityEnd": 'B26',
                        "totalColumn": 'F',
                    },
                    "ACCIDENTES DEL TRAYECTO": {
                        "economicActivityStart": 'B28',
                        "economicActivityEnd": 'B45',
                        "totalColumn": 'F',
                    },
                    "TASA ACCIDENTABILIDAD": {
                        "economicActivityStart": 'B47',
                        "economicActivityEnd": 'B64',
                        "totalColumn": 'F',
                    },
                }
                
                data = {}
                
                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']
                    
                    economicActivities = []
                    totalValues = []
                    
                    for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                        economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value
                        
                        achsValue = economicActivityCell.offset(column=1).value
                        musegValue = economicActivityCell.offset(column=2).value
                        istValue = economicActivityCell.offset(column=3).value
                        totalValue = ws[f"{totalColumn}{row}"].value
                        
                        achsValue = round(achsValue * 100, 1)
                        musegValue = round(musegValue * 100, 1)
                        istValue = round(istValue * 100, 1)
                        totalValue = round(totalValue * 100, 1)
                        
                        
                        
                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "ACHS": achsValue,
                            "MUSEG": musegValue,
                            "IST": istValue,
                        })
                        totalValues.append(totalValue)
                    
                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities
                
                for category, categoryInfo in data.items():
                    print(category)
                    for i, economicActivity in enumerate(categoryInfo['Economic Activities']):
                        print(f"Economic activity: {economicActivity['Economic Activity']}")
                        print(f"ACHS: {economicActivity['ACHS']}%")
                        print(f"MUSEG: {economicActivity['MUSEG']}%")
                        print(f"IST: {economicActivity['IST']}%")
                        print(f"Total: {categoryInfo['Total'][i]}%")
                    print()  

def extractedDataFromExcel_Type2(filePath, extracted_year):
       print(f"archivo procesado: {filePath}")
       
       workbook = openpyxl.load_workbook(filePath)
       #33 NÚMERO PROMEDIO DE DÍAS PERDIDOS POR CADA ACCIDENTES DEL TRABAJO Y DE TRAYECTO SEGÚN ACTIVIDAD ECONÓMICA Y MUTUALIDADES
       #40 NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR
       #no usable'31' numero promedio de dias para todos los archivos bajo 2022. originalmente es la sheet 29 de la function type 1
       sheetsToProcess = ['33', '40', '41','30', '25'] 
       sheetsToProcess2014 = ['34','30', '28']
       sheetsToProcess2015 = ['37', '31', '28']
       #heres where basd on the param extracted_year the logic decides what
       #excel sheet to use    
       if extracted_year == 2014:
           sheetsToProcess = sheetsToProcess2014
       elif extracted_year == 2015:
           sheetsToProcess =  sheetsToProcess2015
       
       for sheet in sheetsToProcess:
           if sheet in workbook.sheetnames:
                wb = workbook[sheet]
                print(f"switched to sheet: {wb.title}")

                if sheet == '40':
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR")
                    processingLogicForSheet40(wb)
                elif sheet == '41':
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO EN MUTUALIDADES E ISL SEGÚN TIPO DE ACCIDENTE, ACTIVIDAD ECONÓMICA Y SEXO")
                    processingLogicForSheet41(wb)
                elif sheet == '25':
                    if extracted_year != 2016:
                        print("processing logic for:NÚMERO DE ACCIDENTES DEL TRABAJO, DE TRAYECTO Y DE ENFERMEDADES PROFESIONALES SEGÚN REGIÓN Y MUTUALIDADES ")
                        processingLogicForSheet25(wb)
                    else:
                        processingLogicForSheet25v2(wb)
                elif sheet == '30':
                    if extracted_year != 2016:
                        print("processing logic for:TASAS DE ACCIDENTABILIDAD POR ACCIDENTES DEL TRABAJO Y DE TRAYECTO SEGÚN ACTIVIDAD ECONÓMICA Y MUTUALIDADES ")
                        processLogicForSheet30(wb)
                    else:
                        print("process logic for TASAS DE ACCIDENTABILIDAD POR ACCIDENTES DEL TRABAJO Y DE TRAYECTO SEGÚN ACTIVIDAD ECONÓMICA Y MUTUALIDADES")
                        processLogicForSheet30In2014(wb)
                        
                elif sheet == '28' and extracted_year in[2014,2015]:
                    print("processing logic for:TASAS DE ACCIDENTABILIDAD POR ACCIDENTES DEL TRABAJO Y DE TRAYECTO, SEGÚN ACTIVIDAD ECONÓMICA Y MUTUAL ")
                    processLogicForSheet30In2014(wb)
                        
                elif sheet == '34' and int(extracted_year) == 2014:
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO, SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR")
                    processingLogicForSheet40In2014(wb)
                elif sheet == '30' and extracted_year == 2014:
                    print("processign logic for: NÚMERO PROMEDIO DE DÍAS PERDIDOS POR CADA ACCIDENTES DEL TRABAJO Y DE TRAYECTO, SEGÚN ACTIVIDAD ECONÓMICA Y MUTUAL")
                    processLogicForSheet31In2014(wb)
                elif sheet == '37' and int(extracted_year) == 2015:
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO, SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR")
                    processingLogicForSheet40In2015(wb)
                elif sheet == '31' and int(extracted_year) == 2015:
                    print("processign logic for: NÚMERO PROMEDIO DE DÍAS PERDIDOS POR CADA ACCIDENTES DEL TRABAJO Y DE TRAYECTO, SEGÚN ACTIVIDAD ECONÓMICA Y MUTUAL ")
                    processigLogicForSheet31In2015(wb)
                else:
                    categoryData = {
                        "ACCIDENTES DEL TRABAJO": {
                                "economicActivityStart":'B10',   
                                "economicActivityEnd": 'B27',
                                "totalColumn": 'F',
                                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        }
                    }        
                    data = {}  
                    for category, categoryInfo in categoryData.items():
                        data[category] = {}
                        
                        economicActivityStart = wb[categoryInfo['economicActivityStart']]
                        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
                        totalColumn = categoryInfo['totalColumn']
                        economicActivities = []

                        for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                            economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                            economicActivity = economic_activity_cell.value
                                    
                            achs_value = economic_activity_cell.offset(column=1).value
                            museg_value = economic_activity_cell.offset(column=2).value
                            ist_value = economic_activity_cell.offset(column=3).value
                            totalValue = wb[f"{totalColumn}{row}"].value 
                            
                            additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                                    
                            economicActivities.append({
                                "Economic Activity": economicActivity,
                                "ACHS": achs_value,
                                "MUSEG": museg_value,
                                "IST": ist_value,
                                "TOTAL": totalValue,
                                "Additional Values": additionalValues,
                            })
                        data[category]['Economic Activities'] = economicActivities

                    for category, categoryInfo in data.items():
                            print(category)
                            for economic_activity in categoryInfo['Economic Activities']:
                                print(f"Economic Activity: {economic_activity['Economic Activity']}")
                                print(f"ACHS: {economic_activity['ACHS']}")
                                print(f"MUSEG: {economic_activity['MUSEG']}")
                                print(f"IST: {economic_activity['IST']}")
                                print(f"Total: {economic_activity['TOTAL']}")
                                
                                #Formato pa valores adicionales / format for the additional vlues
                                additionalValuesLabels = ['ACHS', 'MUSEG', "IST", "TOTAL"]
                                additionalValuesStr = ", ".join([f"{label}: {value}" for label, value in zip(additionalValuesLabels, economic_activity['Additional Values'])])
                                
                                
                                print(additionalValuesStr.replace(", ", ",\n"))
                                print()

def processingLogicForSheet40(wb):
                print(f"custom logic for sheet: {wb.title}")
                categoryData = {
                    "ACCIDENTES DEL TRABAJO": {
                        "economicActivityStart": 'B8',
                        "economicActivityEnd": 'B24',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES DE TRAYECTO":{
                        "economicActivityStart": 'B26',
                        "economicActivityEnd": 'B42',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)":{
                        "economicActivityStart": 'B44',
                        "economicActivityEnd": 'B60',
                        "totalColumn": 'G'
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}

                    economicActivityStart = wb[categoryInfo['economicActivityStart']]
                    economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']

                    economicActivities = []
                    totalValues = []

                    for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
                        economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value

                        achsValue = economicActivityCell.offset(column=1).value
                        musegValue = economicActivityCell.offset(column=2).value
                        istValue = economicActivityCell.offset(column=3).value
                        islValue = economicActivityCell.offset(column=4).value
                        totalValue = wb[f"{totalColumn}{row}"].value

                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "ACHS": achsValue,
                            "MUSEG": musegValue,
                            "IST": istValue,
                            "ISL": islValue
                        })
                        totalValues.append(totalValue)

                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities

                for category, categoryInfo in data.items():
                    print(category)
                    for i, economicActivity in enumerate(categoryInfo['Economic Activities']):
                        print(f"Economic activity: {economicActivity['Economic Activity']}")
                        print(f"ACHS: {economicActivity['ACHS']}")
                        print(f"MUSEG: {economicActivity['MUSEG']}")
                        print(f"IST: {economicActivity['IST']}")
                        print(f"ISL: {economicActivity['ISL']}")
                        print(f"Total: {categoryInfo['Total'][i]}")
                    print()      
def processingLogicForSheet31(wb):
                print()
                categoryData = {
                    "ACCIDENTES DEL TRABAJO": {
                        "mutualInicio": "B8",
                        "mutualFinal": "B11",

                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "mutualInicio": "B13",
                        "mutualFinal": "B16",
                        
                    },
                    "POR ACCIDENTES (TRABAJO + TRAYECTO)": {
                        "mutualInicio": "B18",
                        "mutualFinal": "B21",
                    
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    mutualInicioCell = wb[categoryInfo["mutualInicio"]]
                    mutualFinalCell = wb[categoryInfo["mutualFinal"]]
                    
        
                    
                    mutuales = []

                    for row in range(mutualInicioCell.row, mutualFinalCell.row + 1):
                        mutualInicioCell = wb.cell(row=row, column=mutualInicioCell.column)
                        mutualInicio = mutualInicioCell.value
                        
                        anio2017 = mutualInicioCell.offset(column=1).value
                        anio2018 = mutualInicioCell.offset(column=2).value
                        anio2019 = mutualInicioCell.offset(column=3).value
                        anio2020 = mutualInicioCell.offset(column=4).value
                        anio2021 = mutualInicioCell.offset(column=5).value

                        mutuales.append({
                            "Mutualidades": mutualInicio,
                            "2017": anio2017,
                            "2018": anio2018,
                            "2019": anio2019,
                            "2020": anio2020,
                            "2021": anio2021
                        })
                    data[category]['Mutuales'] = mutuales
                        
                for category, categoryInfo in data.items():
                    print(category)
                    for i, mutualInicio in enumerate(categoryInfo['Mutuales']):
                        print(f"mutuales: {mutualInicio['Mutualidades']}")
                        print(f"2018: {mutualInicio['2017']}")
                        print(f"2019: {mutualInicio['2018']}")
                        print(f"2020: {mutualInicio['2019']}")
                        print(f"2021: {mutualInicio['2020']}")
                        print(f"2022: {mutualInicio['2021']}")
                    print()                   
def processingLogicForSheet40In2014(wb):
    print("custom logic for the page 34 in 2014")
    categoryData = {
        "ACCIDENTES DEL TRABAJO":{
            "economicActivityStart": 'B8',
            "economicActivityEnd": 'B17',
            "totalColumn":'G'
        },
        "ACCIDENTES DEL TRAYECTO": {
            "economicActivityStart": 'B19',
            "economicActivityEnd": 'B28',
            "totalColumn": 'G'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "economicActivityStart": 'B30',
            "economicActivityEnd": 'B39',
            "totalColumn": 'G'
        }
    }
    data = {}
    for category, categoryInfo in categoryData.items():
        data[category] = {}

        economicActivityStart = wb[categoryInfo['economicActivityStart']]
        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
        totalColumn = categoryInfo['totalColumn']

        economicActivities = []
        totalValues = []

        for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
            economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
            economicActivity = economicActivityCell.value

            achsValue = economicActivityCell.offset(column=1).value
            cchc = economicActivityCell.offset(column=2).value
            istValue = economicActivityCell.offset(column=3).value
            islValue = economicActivityCell.offset(column=4).value
            totalValue = wb[f"{totalColumn}{row}"].value

            economicActivities.append({
                "Economic Activity": economicActivity,
                "ACHS": achsValue,
                "CCHC": cchc,
                "IST": istValue,
                "ISL": islValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Economic Activities'] = economicActivities

    for category, categoryInfo in data.items():
        print(category)
        for i, economicActivity in enumerate(categoryInfo['Economic Activities']):
            print(f"Economic activity: {economicActivity['Economic Activity']}")
            print(f"ACHS: {economicActivity['ACHS']}")
            print(f"MUSEG: {economicActivity['CCHC']}")
            print(f"IST: {economicActivity['IST']}")
            print(f"ISL: {economicActivity['ISL']}")
            print(f"Total: {categoryInfo['Total'][i]}")
        print()   
def processingLogicForSheet40In2015(wb):
    categoryData = {
        "ACCIDENTES DEL TRABAJO":{
            "economicActivityStart": 'B8',
            "economicActivityEnd": 'B16',
            "totalColumn": 'G'
        },
        "ACCIDENTES DEL TRAYECTO": {
            "economicActivityStart": 'B18',
            "economicActivityEnd": 'B27',
            "totalColumn": 'G'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "economicActivityStart": 'B29',
            "economicActivityEnd": 'B37',
            "totalColumn": 'G'
        }
    }
    data = {}
    for category, categoryInfo in categoryData.items():
        data[category] = {}

        economicActivityStart = wb[categoryInfo['economicActivityStart']]
        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
        totalColumn = categoryInfo['totalColumn']

        economicActivities = []
        totalValues = []

        for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
            economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
            economicActivity = economicActivityCell.value

            achsValue = economicActivityCell.offset(column=1).value
            cchc = economicActivityCell.offset(column=2).value
            istValue = economicActivityCell.offset(column=3).value
            islValue = economicActivityCell.offset(column=4).value
            totalValue = wb[f"{totalColumn}{row}"].value

            economicActivities.append({
                "Economic Activity": economicActivity,
                "ACHS": achsValue,
                "CCHC": cchc,
                "IST": istValue,
                "ISL": islValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Economic Activities'] = economicActivities

    for category, categoryInfo in data.items():
        print(category)
        for i, economicActivity in enumerate(categoryInfo['Economic Activities']):
            print(f"Economic activity: {economicActivity['Economic Activity']}")
            print(f"ACHS: {economicActivity['ACHS']}")
            print(f"MUSEG: {economicActivity['CCHC']}")
            print(f"IST: {economicActivity['IST']}")
            print(f"ISL: {economicActivity['ISL']}")
            print(f"Total: {categoryInfo['Total'][i]}")
        print()   
def processigLogicForSheet31In2015(wb):
        categoryData = {
            "ACCIDENTES DEL TRABAJO": {
                    "economicActivityStart":'B10',   
                    "economicActivityEnd": 'B27',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES DEL TRAYECTO":{
                    "economicActivityStart": 'B29',
                    "economicActivityEnd": 'B46',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES (TRABAJO + TRAYECTO)": {
                "economicActivityStart": 'B48',
                "economicActivityEnd": 'B65',
                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                "totalColumn": 'F',
            }
        }        
        data = {}  
        for category, categoryInfo in categoryData.items():
            data[category] = {}
            
            economicActivityStart = wb[categoryInfo['economicActivityStart']]
            economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
            totalColumn = categoryInfo['totalColumn']
            economicActivities = []

            for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                economicActivity = economic_activity_cell.value
                        
                achs_value = economic_activity_cell.offset(column=1).value
                cchc_value = economic_activity_cell.offset(column=2).value
                ist_value = economic_activity_cell.offset(column=3).value
                totalValue = wb[f"{totalColumn}{row}"].value 
                
                additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                        
                economicActivities.append({
                    "Economic Activity": economicActivity,
                    "ACHS": achs_value,
                    "CCHC": cchc_value,
                    "IST": ist_value,
                    "TOTAL": totalValue,
                    "Additional Values": additionalValues,
                })
            data[category]['Economic Activities'] = economicActivities

        for category, categoryInfo in data.items():
                print(category)
                for economic_activity in categoryInfo['Economic Activities']:
                    print(f"Economic Activity: {economic_activity['Economic Activity']}")
                    print(f"ACHS: {economic_activity['ACHS']}")
                    print(f"CCHC: {economic_activity['CCHC']}")
                    print(f"IST: {economic_activity['IST']}")
                    print(f"Total: {economic_activity['TOTAL']}")
                    
                    #Formato pa valores adicionales / format for the additional vlues
                    additionalValuesLabels = ['ACHS', 'CCHC', "IST", "TOTAL"]
                    additionalValuesStr = ", ".join([f"{label}: {value}" for label, value in zip(additionalValuesLabels, economic_activity['Additional Values'])])
                    
                    
                    print(additionalValuesStr.replace(", ", ",\n"))
                    print()
def processLogicForSheet31In2014(wb):
        categoryData = {
            "ACCIDENTES DEL TRABAJO": {
                    "economicActivityStart":'B10',   
                    "economicActivityEnd": 'B20',
                    "totalColumn": 'G',
                    "additionalValuesColumns": [ 'H', 'J', 'K', 'L'],
            },
            "ACCIDENTES DEL TRAYECTO":{
                    "economicActivityStart": 'B22',
                    "economicActivityEnd": 'B32',
                    "totalColumn": 'G',
                    "additionalValuesColumns": [  'H', 'J', 'K', 'L'],
            },
            "ACCIDENTES (TRABAJO + TRAYECTO)": {
                "economicActivityStart": 'B34',
                "economicActivityEnd": 'B44',
                "additionalValuesColumns": ['H', 'J', 'K', 'L'],
                "totalColumn": 'G',
            }
        }        
        data = {}  
        for category, categoryInfo in categoryData.items():
            data[category] = {}
            
            economicActivityStart = wb[categoryInfo['economicActivityStart']]
            economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
            totalColumn = categoryInfo['totalColumn']
            economicActivities = []

            for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                economicActivity = economic_activity_cell.value
                        
                achs_value = economic_activity_cell.offset(column=1).value
                cchc_value = economic_activity_cell.offset(column=2).value
                ist_value = economic_activity_cell.offset(column=3).value
                totalValue = wb[f"{totalColumn}{row}"].value 
                
                additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                        
                economicActivities.append({
                    "Economic Activity": economicActivity,
                    "ACHS": achs_value,
                    "CCHC": cchc_value,
                    "IST": ist_value,
                    "TOTAL": totalValue,
                    "Additional Values": additionalValues,
                })
            data[category]['Economic Activities'] = economicActivities

        for category, categoryInfo in data.items():
                print(category)
                for economic_activity in categoryInfo['Economic Activities']:
                    print(f"Economic Activity: {economic_activity['Economic Activity']}")
                    print(f"ACHS: {economic_activity['ACHS']}")
                    print(f"CCHC: {economic_activity['CCHC']}")
                    print(f"IST: {economic_activity['IST']}")
                    print(f"Total: {economic_activity['TOTAL']}")
                    
                    #Formato pa valores adicionales / format for the additional vlues
                    additionalValuesLabels = ['ACHS', 'CCHC', "IST", "TOTAL"]
                    additionalValuesStr = ", ".join([f"{label}: {value}" for label, value in zip(additionalValuesLabels, economic_activity['Additional Values'])])
                    
                    
                    print(additionalValuesStr.replace(", ", ",\n"))
                    print()
def processLogicForSheet29(wb):
                print()
                categoryData = {
                    "ACCIDENTES DEL TRABAJO": {
                        "mutualInicio": "B8",
                        "mutualFinal": "B11",

                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "mutualInicio": "B13",
                        "mutualFinal": "B16",
                        
                    },
                    "POR ACCIDENTES (TRABAJO + TRAYECTO)": {
                        "mutualInicio": "B18",
                        "mutualFinal": "B21",
                    
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    mutualInicioCell = wb[categoryInfo["mutualInicio"]]
                    mutualFinalCell = wb[categoryInfo["mutualFinal"]]
                    
        
                    
                    mutuales = []

                    for row in range(mutualInicioCell.row, mutualFinalCell.row + 1):
                        mutualInicioCell = wb.cell(row=row, column=mutualInicioCell.column)
                        mutualInicio = mutualInicioCell.value
                        
                        anio2017 = mutualInicioCell.offset(column=1).value
                        anio2018 = mutualInicioCell.offset(column=2).value
                        anio2019 = mutualInicioCell.offset(column=3).value
                        anio2020 = mutualInicioCell.offset(column=4).value
                        anio2021 = mutualInicioCell.offset(column=5).value

                        mutuales.append({
                            "Mutualidades": mutualInicio,
                            "2018": anio2017,
                            "2019": anio2018,
                            "2020": anio2019,
                            "2021": anio2020,
                            "2022": anio2021
                        })
                    data[category]['Mutuales'] = mutuales
                        
                for category, categoryInfo in data.items():
                    print(category)
                    for i, mutualInicio in enumerate(categoryInfo['Mutuales']):
                        print(f"mutuales: {mutualInicio['Mutualidades']}")
                        print(f"2018: {mutualInicio['2018']}")
                        print(f"2019: {mutualInicio['2019']}")
                        print(f"2020: {mutualInicio['2020']}")
                        print(f"2021: {mutualInicio['2021']}")
                        print(f"2022: {mutualInicio['2022']}")
                    print()                   
def processingLogicForSheet41(wb):
    print(f"custom logic for sheet: {wb.title}")
    categoryData={
        "ACCIDENTES DEL TRABAJO": {
            "economicActivityStart": 'B8',
            "economicActivityEnd": 'B24',
            "totalColumn": 'E'
        },
        "ACCIDENTES DE TRAYECTO":{
            "economicActivityStart": 'B26',
            "economicActivityEnd": 'B42',
            "totalColumn": 'E'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)":{
            "economicActivityStart": 'B44',
            "economicActivityEnd": 'B60',
            "totalColumn": 'E'
        }
    }
    data = {}

    for category, categoryInfo in categoryData.items():
        data[category] = {}

        economicActivityStart = wb[categoryInfo['economicActivityStart']]
        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
        totalColumn = categoryInfo['totalColumn']

        economicActivities = []
        totalValues = []

        for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
            economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
            economicActivity = economicActivityCell.value

            menValue = economicActivityCell.offset(column=1).value
            womenValue = economicActivityCell.offset(column=2).value
            totalValue = wb[f"{totalColumn}{row}"].value

            economicActivities.append({
                "Economic Activity": economicActivity,
                "MEN": menValue,
                "WOMEN": womenValue,
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Economic Activities'] = economicActivities

    for category, categoryInfo in data.items():
        print(category)
        for i, economicActivity in enumerate(categoryInfo['Economic Activities']):
            print(f"Economic activity: {economicActivity['Economic Activity']}")
            print(f"MEN: {economicActivity['MEN']}")
            print(f"WOMEN: {economicActivity['WOMEN']}")
            print(f"Total: {categoryInfo['Total'][i]}")
        print()      
def processLogicForSheet30(wb):
        categoryData = {
            "ACCIDENTES DEL TRABAJO": {
                    "economicActivityStart":'B10',   
                    "economicActivityEnd": 'B27',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES DEL TRAYECTO":{
                    "economicActivityStart": 'B29',
                    "economicActivityEnd": 'B46',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES (TRABAJO + TRAYECTO)": {
                "economicActivityStart": 'B48',
                "economicActivityEnd": 'B65',
                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                "totalColumn": 'F',
            }
        }        
        data = {}  
        for category, categoryInfo in categoryData.items():
            data[category] = {}
            
            economicActivityStart = wb[categoryInfo['economicActivityStart']]
            economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
            totalColumn = categoryInfo['totalColumn']
            economicActivities = []

            for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                economicActivity = economic_activity_cell.value
                        
                achs_value = economic_activity_cell.offset(column=1).value
                museg_value = economic_activity_cell.offset(column=2).value
                ist_value = economic_activity_cell.offset(column=3).value
                totalValue = wb[f"{totalColumn}{row}"].value 
                
                additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                        
                economicActivities.append({
                    "Economic Activity": economicActivity,
                    "ACHS": achs_value,
                    "MUSEG": museg_value,
                    "IST": ist_value,
                    "TOTAL": totalValue,
                    "Additional Values": additionalValues,
                })
            data[category]['Economic Activities'] = economicActivities

        for category, categoryInfo in data.items():
                print(category)
                for economic_activity in categoryInfo['Economic Activities']:
                    print(f"Economic Activity: {economic_activity['Economic Activity']}")
                    print(f"ACHS: {economic_activity['ACHS']}")
                    print(f"MUSEG: {economic_activity['MUSEG']}")
                    print(f"IST: {economic_activity['IST']}")
                    print(f"Total: {economic_activity['TOTAL']}")
                    
                    #Formato pa valores adicionales / format for the additional vlues
                    additionalValuesLabels = ['ACHS', 'MUSEG', "IST", "TOTAL"]
                    additionalValuesStr = ", ".join([f"{label}: {value}" for label, value in zip(additionalValuesLabels, economic_activity['Additional Values'])])
                    
                    
                    print(additionalValuesStr.replace(", ", ",\n"))
                    print()
def processLogicForSheet30In2014(wb):
                    categoryData = {
                        "TASA ACCIDENTES DEL TRABAJO": {
                                "economicActivityStart":'B10',   
                                "economicActivityEnd": 'B27',
                                "totalColumn": 'F',
                                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        },
                        "TASA ACCIDENTES DEL TRAYECTO": {
                            "economicActivityStart":'B29',   
                            "economicActivityEnd": 'B46',
                            "totalColumn": 'F',
                            "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        },
                        "TASA ACCIDENTABILIDAD": {
                            "economicActivityStart":'B48',   
                            "economicActivityEnd": 'B65',
                            "totalColumn": 'F',
                            "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        }
                    }        
                    data = {}  
                    for category, categoryInfo in categoryData.items():
                        data[category] = {}
                        
                        economicActivityStart = wb[categoryInfo['economicActivityStart']]
                        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
                        totalColumn = categoryInfo['totalColumn']
                        economicActivities = []

                        for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                            economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                            economicActivity = economic_activity_cell.value
                                    
                            achs_value = economic_activity_cell.offset(column=1).value
                            cchc_value = economic_activity_cell.offset(column=2).value
                            ist_value = economic_activity_cell.offset(column=3).value
                            totalValue = wb[f"{totalColumn}{row}"].value 
                            
                            additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                                    
                            economicActivities.append({
                                "Economic Activity": economicActivity,
                                "ACHS": achs_value,
                                "CCHC": cchc_value,
                                "IST": ist_value,
                                "TOTAL": totalValue,
                                "Additional Values": additionalValues,
                            })
                        data[category]['Economic Activities'] = economicActivities

                    for category, categoryInfo in data.items():
                            print(category)
                            for economic_activity in categoryInfo['Economic Activities']:
                                print(f"Economic Activity: {economic_activity['Economic Activity']}")
                                print(f"ACHS: {economic_activity['ACHS']}")
                                print(f"CCHC: {economic_activity['CCHC']}")
                                print(f"IST: {economic_activity['IST']}")
                                print(f"Total: {economic_activity['TOTAL']}")
                                
                                #Formato pa valores adicionales / format for the additional vlues
                                additionalValuesLabels = ['ACHS', 'CCHC', "IST", "TOTAL"]
                                additionalValuesStr = ", ".join([f"{label}: {value}" for label, value in zip(additionalValuesLabels, economic_activity['Additional Values'])])
                                
                                
                                print(additionalValuesStr.replace(", ", ",\n"))
                                print()
def processingLogicForSheet25(wb):
    categoryData = {
        "ACCIDENTES DEL TRABAJO":{
            "regionStart": 'B9',
            "regionEnd": 'B25',
            "totalColumn": 'F'
        },
        "ACCIDENTES DEL TRAYECTO":{
            "regionStart": 'B27',
            "regionEnd": 'B43',
            "totalColumn": 'F'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "regionStart": 'B45',
            "regionEnd": 'B61',
            "totalColumn": 'F'
        },
        "ENFERMEDADES PROFESIONALES": {
            "regionStart": 'B63',
            "regionEnd": 'B79',
            "totalColumn": 'F'
        }
    }
    data = {}

    for category, categoryInfo in categoryData.items():
        data[category] = {}

        regionStart = wb[categoryInfo['regionStart']]
        regionEnd = wb[categoryInfo['regionEnd']]
        totalColumn = categoryInfo['totalColumn']

        regionActivities = []
        totalValues = []

        for row in range (regionStart.row, regionEnd.row + 1):
            regionStartCell = wb.cell(row=row,column=regionStart.column)
            regionActivity = regionStartCell.value

            achsValue = regionStartCell.offset(column=1).value
            musegValue = regionStartCell.offset(column=2).value
            istValue = regionStartCell.offset(column=3).value
            totalValue = wb[f"{totalColumn}{row}"].value

            regionActivities.append({
                "Activity region": regionActivity,
                "ACHS": achsValue,
                "MUSEG": musegValue,
                "IST": istValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Region Activities'] = regionActivities

    for category, categoryInfo in data.items():
        print(category)
        for i, regionActivity in enumerate(categoryInfo['Region Activities']):
            print(f"Activity region: {regionActivity['Activity region']}")
            print(f"ACHS: {regionActivity['ACHS']}")
            print(f"MUSEG: {regionActivity['MUSEG']}")
            print(f"IST: {regionActivity['IST']}")
            print(f"Total: {categoryInfo['Total'][i]}")
        print()   
        
def processingLogicForSheet25v2(wb):
    categoryData = {
        "ACCIDENTES DEL TRABAJO":{
            "regionStart": 'B9',
            "regionEnd": 'B25',
            "totalColumn": 'F'
        },
        "ACCIDENTES DEL TRAYECTO":{
            "regionStart": 'B27',
            "regionEnd": 'B43',
            "totalColumn": 'F'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "regionStart": 'B45',
            "regionEnd": 'B61',
            "totalColumn": 'F'
        },
        "ENFERMEDADES PROFESIONALES": {
            "regionStart": 'B63',
            "regionEnd": 'B79',
            "totalColumn": 'F'
        }
    }
    data = {}

    for category, categoryInfo in categoryData.items():
        data[category] = {}

        regionStart = wb[categoryInfo['regionStart']]
        regionEnd = wb[categoryInfo['regionEnd']]
        totalColumn = categoryInfo['totalColumn']

        regionActivities = []
        totalValues = []

        for row in range (regionStart.row, regionEnd.row + 1):
            regionStartCell = wb.cell(row=row,column=regionStart.column)
            regionActivity = regionStartCell.value

            achsValue = regionStartCell.offset(column=1).value
            musegValue = regionStartCell.offset(column=2).value
            istValue = regionStartCell.offset(column=3).value
            totalValue = wb[f"{totalColumn}{row}"].value

            regionActivities.append({
                "Activity region": regionActivity,
                "ACHS": achsValue,
                "MUSEG": musegValue,
                "IST": istValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Region Activities'] = regionActivities

    for category, categoryInfo in data.items():
        print(category)
        for i, regionActivity in enumerate(categoryInfo['Region Activities']):
            print(f"Activity region: {regionActivity['Activity region']}")
            print(f"ACHS: {regionActivity['ACHS']}")
            print(f"MUSEG: {regionActivity['MUSEG']}")
            print(f"IST: {regionActivity['IST']}")
            print(f"Total: {categoryInfo['Total'][i]}")
        print()
        
def getFileYear(fileName):
    decodedFileName = fileName.encode('utf-8').decode('utf-8')
    cleanedFileName = unidecode.unidecode(decodedFileName)
    match = re.search(r'\d{4}', cleanedFileName)
    if match:
        year = int(match.group())
        print(f"Extracted year: {year}")
        return year
    return None

def is_2022_file(fileName):
    
    cleanedFileName = unidecode.unidecode(fileName)
    pattern = re.compile(r'2022', re.IGNORECASE)
    result =  re.search(pattern, cleanedFileName) is not None
    print(f"Original File Name: {fileName}")
    print(f"Cleaned File Name: {cleanedFileName}")
    print(f"Checking if {fileName} is a 2022 file: {result}")
    return result

if __name__ == "__main__":
    webPageUrl = "https://www.suseso.cl/608/w3-propertyvalue-10364.html"# Replace with your URL
    downloadDir = os.path.join(os.getcwd(), "downloadedFiles")
    
    
    downloadedFiles = downloadExcelFiles(webPageUrl, downloadDir)
    downloadedFiles.sort(key=lambda x: getFileYear(x[0]) or 0, reverse=True)

    
    lastProcessedFileYear = None
    
    
    for fileInfo in downloadedFiles:
        title, filePath = fileInfo
        if is_2022_file(title):
            print(f"Identified as 2022. Processing with type1")
            extractedDataFromExcel_Type1(filePath)
            lastProcessedFileYear = 2022
        else:
            year = getFileYear(title)
            if lastProcessedFileYear is None or year == lastProcessedFileYear - 1:
                extractedDataFromExcel_Type2(filePath, year)
                lastProcessedFileYear = year
            else:
                print(f"Skipping file: {title} (Unexpected year)")
    
    
    """
    for title, filePath in downloadedFiles:
        if is_2022_file(title):
            print(f"Identified as 2022. Processing with type1")
            extractedDataFromExcel_Type1(filePath)
            lastProcessedFileYear = 2022
        else:
            year = getFileYear(title)
            if lastProcessedFileYear is None or year == lastProcessedFileYear - 1:
                extractedDataFromExcel_Type2(filePath, year)
                lastProcessedFileYear = year
            else:
                print(f"Skipping file: {title} (Unexpected year)")
    """