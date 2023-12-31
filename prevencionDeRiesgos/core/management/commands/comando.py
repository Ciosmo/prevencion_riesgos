import requests
import os
import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from django.core.management.base import BaseCommand
from core.models import EconomicActivity, Category, Mutualidad, Tasa_Eco_Act, Sexo

class Command(BaseCommand):
    help = 'Extraer datos de la pagina SUSESO y guardarlos en la BD'

    def handle(self, *args, **options):
        # Lógica para descargar el archivo y extraer datos
        web_page_url = "https://www.suseso.cl/608/w3-article-708568.html"

        response = requests.get(web_page_url)

        soup = BeautifulSoup(response.text, 'lxml') 

        downloadDir = os.path.join(os.getcwd(), "downloadedFiles")
        os.makedirs(downloadDir, exist_ok=True)
        try:
            for a_tag in soup.find_all('a'):
                if a_tag.get("href") == "articles-708568_archivo_01.xlsx" and a_tag.get("title") == "Ir a EstadÃ­sticas de la Seguridad Social 2022":
                
                    downloadUrl = "https://www.suseso.cl/608/articles-708568_archivo_01.xlsx"

                    fileName = os.path.basename(downloadUrl)
                    localPath = os.path.join(downloadDir, fileName)
                
                    fileResponse = requests.get(downloadUrl)

                    if fileResponse.status_code == 200:
                        with open(localPath, "wb") as localFile:
                            localFile.write(fileResponse.content)
                        print(f"{localPath} descargado correctamente")
                        #testing
                        workbookv2 = openpyxl.load_workbook(localPath)
                        sheetsToProcess = ['31', '29', '38', '39']
                        for wantedSheet in sheetsToProcess:
                            if wantedSheet in workbookv2.sheetnames:
                                ws = workbookv2[wantedSheet]
                                if wantedSheet == '31':       
                                    print(f"switched to sheet: {ws.title}") 
                
                            
                                    categoryData = {
                                            "ACCIDENTES DEL TRABAJO": {
                                                "economicActivityStart": 'B9',
                                                "economicActivityEnd": 'B26',
                                                "totalColumn": 'F',
                                            },
                                            "ACCIDENTES DE TRAYECTO": {
                                                "economicActivityStart": 'B28',
                                                "economicActivityEnd": 'B45',  # Updated to 'B45'
                                                "totalColumn": 'F',
                                            },
                                            "ACCIDENTES (TRABAJO + TRAYECTO)": {
                                                "economicActivityStart": 'B48',  # Updated to 'B48'
                                                "economicActivityEnd": 'B64',
                                                "totalColumn": 'F',
                                            },
                                    }

                                    data = {}

                                    for category, categoryInfo in categoryData.items():
                                        data[category] = {}
                                        
                                        economicActivityStart = ws[categoryInfo['economicActivityStart']]
                                        economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                                        total_column = categoryInfo['totalColumn']
                                        
                                        economicActivities = []
                                        total_values = []
                        
                                        for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                                            economic_activity_cell = ws.cell(row=row, column=economicActivityStart.column)
                                            economic_activity = economic_activity_cell.value
                                            
                                            achs_value = economic_activity_cell.offset(column=1).value
                                            museg_value = economic_activity_cell.offset(column=2).value
                                            ist_value = economic_activity_cell.offset(column=3).value
                                            total_value = ws[f"{total_column}{row}"].value  # Fetch the "Total" from the specified column
                                            
                                            economicActivities.append({
                                                "Economic Activity": economic_activity,
                                                "ACHS": achs_value,
                                                "MUSEG": museg_value,
                                                "IST": ist_value,
                                            })
                                            total_values.append(total_value)

                                        data[category]['Total'] = total_values  # Store all "Total" values for each economic activity
                                        data[category]['Economic Activities'] = economicActivities

                                    for category, categoryInfo in data.items():
                                
                                        category_instance, created = Category.objects.get_or_create(name=category)

                                        for i, economic_activity in enumerate(categoryInfo['Economic Activities']):
                                            economic_activity_instance = EconomicActivity(
                                                category=category_instance,
                                                name=economic_activity['Economic Activity'],
                                                achs=economic_activity['ACHS'],
                                                museg=economic_activity['MUSEG'],
                                                ist=economic_activity['IST'],
                                                total=total_value
                                            )
                                            economic_activity_instance.save()

                                if wantedSheet == '29':
                                    print(f"switched to sheet: {ws.title}")
                                    categoryData = {
                                        "ACCIDENTES DEL TRABAJO": {
                                            "mutualInicio": "B8",
                                            "mutualFinal": "B11",
                    
                                        },
                                        "ACCIDENTES DE TRAYECTO": {
                                            "mutualInicio": "B13",
                                            "mutualFinal": "B16",
                                        
                                        },
                                        "ACCIDENTES (TRABAJO + TRAYECTO)": {
                                            "mutualInicio": "B18",
                                            "mutualFinal": "B21",
                                        
                                        }
                                    }
                                    data = {}

                                    for category, categoryInfo in categoryData.items():
                                        data[category] = {}
                                        
                                        mutualInicioCell = ws[categoryInfo["mutualInicio"]]
                                        mutualFinalCell = ws[categoryInfo["mutualFinal"]]
                                        
                            
                                        
                                        mutuales = []
        
                                        for row in range(mutualInicioCell.row, mutualFinalCell.row + 1):
                                            mutualInicioCell = ws.cell(row=row, column=mutualInicioCell.column)
                                            mutualInicio = mutualInicioCell.value
                                            
                                            anio2018 = mutualInicioCell.offset(column=1).value
                                            anio2019 = mutualInicioCell.offset(column=2).value
                                            anio2020 = mutualInicioCell.offset(column=3).value
                                            anio2021 = mutualInicioCell.offset(column=4).value
                                            anio2022 = mutualInicioCell.offset(column=5).value

                                            mutuales.append({
                                                "Mutualidades": mutualInicio,
                                                "2018": anio2018,
                                                "2019": anio2019,
                                                "2020": anio2020,
                                                "2021": anio2021,
                                                "2022": anio2022
                                            })
                                        data[category]['Mutuales'] = mutuales
                                            
                                    for category, categoryInfo in data.items():

                                        category_instance, created = Category.objects.get_or_create(name=category)
                                        
                                        for mutualidad_info in categoryInfo['Mutuales']:
                                            mutualidad_instance = Mutualidad(
                                                category=category_instance,
                                                mutual=mutualidad_info['Mutualidades'],
                                                anio2018=mutualidad_info['2018'],
                                                anio2019=mutualidad_info['2019'],
                                                anio2020=mutualidad_info['2020'],
                                                anio2021=mutualidad_info['2021'],
                                                anio2022=mutualidad_info['2022']
                                            )
                                            mutualidad_instance.save()
                                            
                                if wantedSheet == '38':
                                    print(f"switched to sheet: {ws.title}")
                                    
                                    categoryData = {
                                        "ACCIDENTES DEL TRABAJO":{
                                            "economicActivityStart": 'B8',
                                            "economicActivityEnd": 'B24',
                                            "totalColumn": 'G'
                                        },
                                        "ACCIDENTES DE TRAYECTO": {
                                            "economicActivityStart": 'B26',
                                            "economicActivityEnd":'B42',
                                            "totalColumn": 'G'
                                        },
                                        "ACCIDENTES (TRABAJO + TRAYECTO)":{
                                            "economicActivityStart":'B44',
                                            "economicActivityEnd": 'B60',
                                            "totalColumn": 'G'
                                        }       
                                    }
                                    
                                    data = {}
                                    
                                    for category, categoryInfo in categoryData.items():
                                        data[category] = {}
                                        
                                        economicActivityStart = ws[categoryInfo['economicActivityStart']]
                                        economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                                        totalColumn = categoryInfo['totalColumn']  
                                        
                                        economicActivities = []
                                        totalValues = []
                                        
                                        for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
                                            economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                                            economicActivity = economicActivityCell.value
                                            
                                            achsValue = economicActivityCell.offset(column=1).value
                                            musegValue = economicActivityCell.offset(column=2).value
                                            istValue = economicActivityCell.offset(column=3).value
                                            islValue = economicActivityCell.offset(column=4).value
                                            totalValue = ws[f"{totalColumn}{row}"].value
                                            
                                            economicActivities.append({
                                                "Economic Activity": economicActivity,
                                                "ACHS": achsValue,
                                                "MUSEG": musegValue,
                                                "IST": istValue,
                                                "ISL": islValue
                                            })
                                            totalValues.append(totalValue)
                                            
                                        data[category]['Total'] = totalValues
                                        data[category]['Economic Activities'] = economicActivities
                                    for category, categoryInfo in data.items():
                                
                                        category_instance, created = Category.objects.get_or_create(name=category)

                                        for i, economic_activity in enumerate(categoryInfo['Economic Activities']):
                                            economic_activity_instance = Tasa_Eco_Act(
                                                category=category_instance,
                                                name=economic_activity['Economic Activity'],
                                                achs=economic_activity['ACHS'],
                                                museg=economic_activity['MUSEG'],
                                                ist=economic_activity['IST'],
                                                total=total_value
                                            )
                                            economic_activity_instance.save()

                                if wantedSheet == '39':
                                    print(f"switched to sheet: {ws.title}")
                                    categoryData = {
                                        "ACCIDENTES DEL TRABAJO": {
                                            "economicActivityStart": 'B8',
                                            "economicActivityEnd": 'B24',
                                            "totalColumn": 'E'
                                        },
                                        "ACCIDENTES DE TRAYECTO": {
                                            "economicActivityStart": 'B26',
                                            "economicActivityEnd": 'B42',
                                            "totalColumn": 'E'
                                        },
                                        "ACCIDENTES (TRABAJO + TRAYECTO)":{
                                            "economicActivityStart": 'B44',
                                            "economicActivityEnd": 'B60',
                                            "totalColumn": 'E'
                                        }
                                    }
                                    
                                    data = {}
                                    
                                    for category, categoryInfo in categoryData.items():
                                        
                                        data[category] = {}
                                        
                                        economicActivityStart = ws[categoryInfo['economicActivityStart']]
                                        economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                                        totalColumn = categoryInfo['totalColumn']
                                        economicActivities = []
                                        totalValues = []
                                        
                                        for row in range (economicActivityStart.row, economicActivityEnd.row+1):
                                            economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                                            economicActivity = economicActivityCell.value
                                            
                                            menValues = economicActivityCell.offset(column=1).value
                                            womenValue = economicActivityCell.offset(column=2).value
                                            totalValue = ws[f"{totalColumn}{row}"].value
                                            
                                            economicActivities.append({
                                                "Economic Activity": economicActivity,
                                                "Men": menValues,
                                                "Women": womenValue
                                            })
                                            totalValues.append(totalValue)
                                        data[category]['Total'] = totalValues
                                        data[category]['Economic Activities'] = economicActivities

                                    for category, categoryInfo in data.items():
                                
                                        category_instance, created = Category.objects.get_or_create(name=category)

                                        for i, economic_activity in enumerate(categoryInfo['Economic Activities']):
                                            economic_activity_instance = Sexo(
                                                category=category_instance,
                                                name=economic_activity['Economic Activity'],
                                                men=economic_activity['Men'],
                                                women=economic_activity['Women'],
                                                total=total_value
                                            )
                                            economic_activity_instance.save()
                                            
                                            
        except requests.exceptions.RequestException as e:
            print(f"An error ocurrred while downloading the file: {e}")