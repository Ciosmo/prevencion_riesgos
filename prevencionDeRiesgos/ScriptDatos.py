import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'prevencionDeRiesgos.settings')

# Configurar Django
django.setup()

import requests
import os
import re
import openpyxl
import unidecode

from bs4 import BeautifulSoup
from urllib.parse import urljoin
from unidecode import unidecode

from core.models import DiasxActividad, DiasxMut, TasaxAct, AccidentesxSexo, PorcentajexAct, FallecidosxAct, FallecidosxSexo, AccidentesxRegion, Year, Category, EconomicActivity, Mutualidad, Region
"""
1) Identificar correctamente el elemento(s) que contienen 
informacion deseada
   Elemento general: 
    <div id="article_i__ss_ar_articulo_archivos_1" class="recuadros download-attr margen-abajo-md"><div class="recuadro format-pdf externo cid-507 aid-708568 binary-archivo_01 format-xlsx media"><a href="articles-708568_archivo_01.xlsx" title="Ir a Estadísticas de la Seguridad Social 2022" download="Estadísticas de la Seguridad Social 2022.pdf">Estadísticas de la Seguridad Social 2022</a></div></div>
 Especificamete:
    div que contiene el link de descarga > article_i__ss_ar_articulo_archivos_1
    <a href="articles-708568_archivo_01.xlsx" title="Ir a Estadísticas de la Seguridad Social 2022" download="Estadísticas de la Seguridad Social 2022.pdf">Estadísticas de la Seguridad Social 2022</a>

"""
def downloadExcelFiles(url, downloadDir):
    resp = requests.get(url)
    soup = BeautifulSoup(resp.text, 'html.parser')
    
    os.makedirs(downloadDir, exist_ok=True)
    
     # Find the table with id 'tabular_generica'
    table = soup.find('table', {'id': 'tabular_generica'})
    
    downloadedFiles = []
    
    if table:
        # Find all rows in the table
        rows = table.find_all('tr')
        for row in rows:
            # Find all cells in the row
            cells = row.find_all(['td', 'th'])
            for cell in cells:
                # Check if the cell contains a link
                link = cell.find('a', href=True)
                if link:
                    downloadUrl = link['href']
                    title = link.get('title', '')
                    if downloadUrl.endswith('.xlsx') and 'Seguridad Social' in title:
                        fullUrl = urljoin(url, downloadUrl)
                        
                        newName = f"{title.replace(' ', '_')}download.xlsx"
                        filePath = os.path.join(downloadDir, newName)
                        
                        if filePath not in [file[1] for file in downloadedFiles]:    
                            with open(filePath, 'wb') as file:
                                file.write(requests.get(fullUrl).content)
                            downloadedFiles.append((title,filePath))
                            print(f"Download: {newName}")
    return downloadedFiles          

year_instances = {}
category_instances = {}
economic_activities_instances = {}
mutualidades_instances = {}
tasas_instances = {}
    
def extractedDataFromExcel_Type1(filePath):  
    wb = openpyxl.load_workbook(filePath)
    sheetsToProcess = ['31', '29', '38', '39', '28']
    for wantedSheet in sheetsToProcess:
        if  wantedSheet in wb.sheetnames:
            ws = wb[wantedSheet]
            if wantedSheet == '31':       
                print(f"switched to sheet: {ws.title}")             
                categoryData = {
                        "ACCIDENTES DE TRABAJO": {
                            "economicActivityStart": 'B9',
                            "economicActivityEnd": 'B26',
                            "totalColumn": 'F',
                        },
                        "ACCIDENTES DE TRAYECTO": {
                            "economicActivityStart": 'B28',
                            "economicActivityEnd": 'B45',  # Updated to 'B45'
                            "totalColumn": 'F',
                        },
                        "ACCIDENTES (TRABAJO + TRAYECTO)": {
                            "economicActivityStart": 'B48',  # Updated to 'B48'
                            "economicActivityEnd": 'B64',
                            "totalColumn": 'F',
                        },
                }

                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    total_column = categoryInfo['totalColumn']
                    
                    economicActivities = []
                    total_values = []

                    for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                        economic_activity_cell = ws.cell(row=row, column=economicActivityStart.column)
                        economic_activity = economic_activity_cell.value
                        
                        achs_value = economic_activity_cell.offset(column=1).value
                        museg_value = economic_activity_cell.offset(column=2).value
                        ist_value = economic_activity_cell.offset(column=3).value
                        total_value = ws[f"{total_column}{row}"].value  # Fetch the "Total" from the specified column
                        
                        economicActivities.append({
                            "Economic Activity": economic_activity,
                            "ACHS": achs_value,
                            "MUSEG": museg_value,
                            "IST": ist_value,
                        })
                        total_values.append(total_value)

                    data[category]['Total'] = total_values  # Store all "Total" values for each economic activity
                    data[category]['Economic Activities'] = economicActivities

                for category, categoryInfo in data.items():
                    for economic_activity_info in categoryInfo['Economic Activities']:
                        activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = economic_activities_instances.get(activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                            economic_activities_instances[activity_name] = economic_activity_instance

                for category, categoryInfo in data.items():
                    category_instance, created = Category.objects.get_or_create(name=category)

                    for i, economic_activity in enumerate(categoryInfo['Economic Activities']):
                        activity_name = economic_activity['Economic Activity']
                        economic_activity_instance = economic_activities_instances[activity_name]

                        # Crear y guardar la instancia de DiasxActividad
                        diasxactividad_instance = DiasxActividad(
                            category=category_instance,
                            EconomicActivity=economic_activity_instance,
                            achs=economic_activity['ACHS'],
                            museg=economic_activity['MUSEG'],
                            ist=economic_activity['IST'],
                            total=total_value
                        )
                        diasxactividad_instance.save()

            if wantedSheet == '29':
                print(f"switched to sheet: {ws.title}")
                categoryData = {
                    "ACCIDENTES DE TRABAJO": {
                        "mutualInicio": "B8",
                        "mutualFinal": "B11",

                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "mutualInicio": "B13",
                        "mutualFinal": "B16",
                        
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)": {
                        "mutualInicio": "B18",
                        "mutualFinal": "B21",
                    
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    mutualInicioCell = ws[categoryInfo["mutualInicio"]]
                    mutualFinalCell = ws[categoryInfo["mutualFinal"]]
                    
        
                    
                    mutuales = []

                    for row in range(mutualInicioCell.row, mutualFinalCell.row + 1):
                        mutualInicioCell = ws.cell(row=row, column=mutualInicioCell.column)
                        mutualInicio = mutualInicioCell.value
                        
                        anio2018 = mutualInicioCell.offset(column=1).value
                        anio2019 = mutualInicioCell.offset(column=2).value
                        anio2020 = mutualInicioCell.offset(column=3).value
                        anio2021 = mutualInicioCell.offset(column=4).value
                        anio2022 = mutualInicioCell.offset(column=5).value

                        mutuales.append({
                            "Mutualidades": mutualInicio,
                            "2018": anio2018,
                            "2019": anio2019,
                            "2020": anio2020,
                            "2021": anio2021,
                            "2022": anio2022
                        })
                    data[category]['Mutuales'] = mutuales
                        
                for category, categoryInfo in data.items():
                    for mutualidad_info in categoryInfo['Mutuales']:
                        mutualidad_name = mutualidad_info['Mutualidades']

                        # Verificar si ya existe una instancia con el mismo nombre
                        mutualidad_instance = mutualidades_instances.get(mutualidad_name)
                        if not mutualidad_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            mutualidad_instance, created = Mutualidad.objects.get_or_create(mutualidad_name=mutualidad_name)
                            mutualidades_instances[mutualidad_name] = mutualidad_instance


                for category, categoryInfo in data.items():
                    category_instance, created = Category.objects.get_or_create(name=category)

                    for mutualidad_info in categoryInfo['Mutuales']:
                        mutualidad_name = mutualidad_info['Mutualidades']
                        mutualidad_instance = mutualidades_instances[mutualidad_name]

                        # Crear y guardar la instancia de DiasxMut
                        diasxmut_instance = DiasxMut(
                            category=category_instance,
                            mutual=mutualidad_instance,
                            anio2018=mutualidad_info['2018'],
                            anio2019=mutualidad_info['2019'],
                            anio2020=mutualidad_info['2020'],
                            anio2021=mutualidad_info['2021'],
                            anio2022=mutualidad_info['2022']
                        )
                        diasxmut_instance.save()

            if wantedSheet == '38':
                print(f"switched to sheet: {ws.title}")
                
                categoryData = {
                    "ACCIDENTES DE TRABAJO":{
                        "economicActivityStart": 'B8',
                        "economicActivityEnd": 'B24',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "economicActivityStart": 'B26',
                        "economicActivityEnd":'B42',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)":{
                        "economicActivityStart":'B44',
                        "economicActivityEnd": 'B60',
                        "totalColumn": 'G'
                    }       
                }
                    
                data = {}
                
                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']  
                    
                    economicActivities = []
                    totalValues = []
                    
                    for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
                        economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value
                        
                        achsValue = economicActivityCell.offset(column=1).value
                        musegValue = economicActivityCell.offset(column=2).value
                        istValue = economicActivityCell.offset(column=3).value
                        islValue = economicActivityCell.offset(column=4).value
                        totalValue = ws[f"{totalColumn}{row}"].value
                        
                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "ACHS": achsValue,
                            "MUSEG": musegValue,
                            "IST": istValue,
                            "ISL": islValue
                        })
                        totalValues.append(totalValue)
                        
                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities

                for category, categoryInfo in data.items():
                    for economic_activity_info in categoryInfo['Economic Activities']:
                        activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = economic_activities_instances.get(activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                            economic_activities_instances[activity_name] = economic_activity_instance

                
                for category, categoryInfo in data.items():
                    category_instance, created = Category.objects.get_or_create(name=category)

                    for economic_activity_info in categoryInfo['Economic Activities']:
                        economic_activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = tasas_instances.get(economic_activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=economic_activity_name)
                            tasas_instances[economic_activity_name] = economic_activity_instance

                        # Crear y guardar la instancia de TasaxAct
                        tasaxAct_instance = TasaxAct(
                            category=category_instance,
                            EconomicActivity=economic_activity_instance,
                            achs=economic_activity_info['ACHS'],
                            museg=economic_activity_info['MUSEG'],
                            ist=economic_activity_info['IST'],
                            total=total_value
                        )
                        tasaxAct_instance.save()

            if wantedSheet == '39':
                print(f"switched to sheet: {ws.title}")
                categoryData = {
                    "ACCIDENTES DE TRABAJO": {
                        "economicActivityStart": 'B8',
                        "economicActivityEnd": 'B24',
                        "totalColumn": 'E'
                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "economicActivityStart": 'B26',
                        "economicActivityEnd": 'B42',
                        "totalColumn": 'E'
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)":{
                        "economicActivityStart": 'B44',
                        "economicActivityEnd": 'B60',
                        "totalColumn": 'E'
                    }
                }
                
                data = {}
                
                for category, categoryInfo in categoryData.items():
                    
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']
                    economicActivities = []
                    totalValues = []
                    
                    for row in range (economicActivityStart.row, economicActivityEnd.row+1):
                        economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value
                        
                        menValues = economicActivityCell.offset(column=1).value
                        womenValue = economicActivityCell.offset(column=2).value
                        totalValue = ws[f"{totalColumn}{row}"].value
                        
                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "Men": menValues,
                            "Women": womenValue
                        })
                        totalValues.append(totalValue)
                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities

                for category, categoryInfo in data.items():
                    for economic_activity_info in categoryInfo['Economic Activities']:
                        activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = economic_activities_instances.get(activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                            economic_activities_instances[activity_name] = economic_activity_instance

                sexos_instances = {}
                for category, categoryInfo in data.items():
                    category_instance, created = Category.objects.get_or_create(name=category)

                    for economic_activity_info in categoryInfo['Economic Activities']:
                        economic_activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = sexos_instances.get(economic_activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=economic_activity_name)
                            sexos_instances[economic_activity_name] = economic_activity_instance

                        # Crear y guardar la instancia de AccidentesxSexo
                        accidentesx_sexo_instance = AccidentesxSexo(
                            category=category_instance,
                            EconomicActivity=economic_activity_instance,
                            men=economic_activity_info['Men'],
                            women=economic_activity_info['Women'],
                            total=total_value
                        )
                        accidentesx_sexo_instance.save()
                    
            if wantedSheet == '28':
                print(f"switched to sheet: {ws.title}")
                categoryData = {
                    "ACCIDENTES DE TRABAJO": {
                        "economicActivityStart": 'B9',
                        "economicActivityEnd": 'B26',
                        "totalColumn": 'F',
                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "economicActivityStart": 'B28',
                        "economicActivityEnd": 'B45',
                        "totalColumn": 'F',
                    },
                    "TASA ACCIDENTABILIDAD": {
                        "economicActivityStart": 'B47',
                        "economicActivityEnd": 'B64',
                        "totalColumn": 'F',
                    },
                }
                
                data = {}
                
                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    economicActivityStart = ws[categoryInfo['economicActivityStart']]
                    economicActivityEnd = ws[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']
                    
                    economicActivities = []
                    totalValues = []
                    
                    for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                        economicActivityCell = ws.cell(row=row, column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value
                        
                        achsValue = economicActivityCell.offset(column=1).value
                        musegValue = economicActivityCell.offset(column=2).value
                        istValue = economicActivityCell.offset(column=3).value
                        totalValue = ws[f"{totalColumn}{row}"].value
                        
                        achsValue = round(achsValue * 100, 1)
                        musegValue = round(musegValue * 100, 1)
                        istValue = round(istValue * 100, 1)
                        totalValue = round(totalValue * 100, 1)
                        
                        
                        
                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "ACHS": achsValue,
                            "MUSEG": musegValue,
                            "IST": istValue,
                        })
                        totalValues.append(totalValue)
                    
                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities

                for category, categoryInfo in data.items():
                    for economic_activity_info in categoryInfo['Economic Activities']:
                        activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = economic_activities_instances.get(activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                            economic_activities_instances[activity_name] = economic_activity_instance

                porcentaje_instances = {}
                for category, categoryInfo in data.items():
                    category_instance, created = Category.objects.get_or_create(name=category)

                    for economic_activity_info in categoryInfo['Economic Activities']:
                        economic_activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = porcentaje_instances.get(economic_activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=economic_activity_name)
                            porcentaje_instances[economic_activity_name] = economic_activity_instance

                        # Crear y guardar la instancia de PorcentajexAct
                        porcentajexAct_instance = PorcentajexAct(
                            category=category_instance,
                            EconomicActivity=economic_activity_instance,
                            achs=economic_activity_info['ACHS'],
                            museg=economic_activity_info['MUSEG'],
                            ist=economic_activity_info['IST'],
                            total=total_value
                        )
                        porcentajexAct_instance.save()

def extractedDataFromExcel_Type2(filePath, extracted_year):
       print(f"archivo procesado: {filePath}")
       
       workbook = openpyxl.load_workbook(filePath)
       #33 NÚMERO PROMEDIO DE DÍAS PERDIDOS POR CADA ACCIDENTES DEL TRABAJO Y DE TRAYECTO SEGÚN ACTIVIDAD ECONÓMICA Y MUTUALIDADES
       #40 NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR
       #no usable'31' numero promedio de dias para todos los archivos bajo 2022. originalmente es la sheet 29 de la function type 1
       sheetsToProcess = ['33', '40', '41','30', '25'] 
       sheetsToProcess2014 = ['34','30', '28']
       sheetsToProcess2015 = ['37', '31', '28']
       #heres where basd on the param extracted_year the logic decides what
       #excel sheet to use    
       if extracted_year == 2014:
           sheetsToProcess = sheetsToProcess2014
       elif extracted_year == 2015:
           sheetsToProcess =  sheetsToProcess2015
       
       for sheet in sheetsToProcess:
           if sheet in workbook.sheetnames:
                wb = workbook[sheet]
                print(f"switched to sheet: {wb.title}")

                if sheet == '40': #la pagina 40 es la misma para los años entre 2016 a 2021 LISTO
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR")
                    processingLogicForSheet40(wb, extracted_year) 
                elif sheet == '41': #la pagina 41 es la misma para los años entre 2016 a 2021 LISTO
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO EN MUTUALIDADES E ISL SEGÚN TIPO DE ACCIDENTE, ACTIVIDAD ECONÓMICA Y SEXO")
                    processingLogicForSheet41(wb, extracted_year)
                elif sheet == '25': #la pagina 25 es la misma para los años entre 2017 a 2021 LISTO
                    if extracted_year != 2016:
                        print("processing logic for:NÚMERO DE ACCIDENTES DEL TRABAJO, DE TRAYECTO Y DE ENFERMEDADES PROFESIONALES SEGÚN REGIÓN Y MUTUALIDADES ")
                        processingLogicForSheet25(wb, extracted_year)
                    else:
                        processingLogicForSheet25v2(wb, extracted_year) #esta se usa para el año 2016 LISTO
                elif sheet == '30':
                    if extracted_year != 2016: #cada archivo trae datos del año anterior y la del año del archivo LISTO PERO SOLO VALORES ADICIONALES
                        print("processing logic for:TASAS DE ACCIDENTABILIDAD POR ACCIDENTES DEL TRABAJO Y DE TRAYECTO SEGÚN ACTIVIDAD ECONÓMICA Y MUTUALIDADES ")
                        processLogicForSheet30(wb, extracted_year)
                    else:
                        print("process logic for TASAS DE ACCIDENTABILIDAD POR ACCIDENTES DEL TRABAJO Y DE TRAYECTO SEGÚN ACTIVIDAD ECONÓMICA Y MUTUALIDADES")
                        processLogicForSheet30In2014(wb, extracted_year) #es lo mismo que la anterior, pero cambia la mutulidad museg por cchc LISTO PERO SOLO VALORES ADICIONALES
                        
                elif sheet == '28' and extracted_year in[2014,2015]: #la misma logica pero para 2014 y 2015 LISTO PERO SOLO VALORES ADICIONALES
                    print("processing logic for:TASAS DE ACCIDENTABILIDAD POR ACCIDENTES DEL TRABAJO Y DE TRAYECTO, SEGÚN ACTIVIDAD ECONÓMICA Y MUTUAL ")
                    processLogicForSheet30In2014(wb, extracted_year)
                        
                elif sheet == '34' and int(extracted_year) == 2014: #es igual a la 40 pero para 2014 LISTO
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO, SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR")
                    processingLogicForSheet40In2014(wb, extracted_year) 
                elif sheet == '30' and extracted_year == 2014: #tiene info del 2013 y 2014 LISTO SOLO DATOS ADICIONALES
                    print("processign logic for: NÚMERO PROMEDIO DE DÍAS PERDIDOS POR CADA ACCIDENTES DEL TRABAJO Y DE TRAYECTO, SEGÚN ACTIVIDAD ECONÓMICA Y MUTUAL")
                    processLogicForSheet31In2014(wb, extracted_year)
                elif sheet == '37' and int(extracted_year) == 2015: #solo 2015 LISTO
                    print("processing logic for: NÚMERO DE FALLECIDOS POR ACCIDENTES DEL TRABAJO, SEGÚN TIPO DE ACCIDENTE,  ACTIVIDAD ECONÓMICA Y ORGANISMO ADMINISTRADOR")
                    processingLogicForSheet40In2015(wb, extracted_year)
                elif sheet == '31' and int(extracted_year) == 2015: #tiene info del 2014 y 2015 LISTO SOLO DATOS ADICIONALES
                    print("processign logic for: NÚMERO PROMEDIO DE DÍAS PERDIDOS POR CADA ACCIDENTES DEL TRABAJO Y DE TRAYECTO, SEGÚN ACTIVIDAD ECONÓMICA Y MUTUAL ")
                    processigLogicForSheet31In2015(wb, extracted_year)
                else: # pagina 33 NÚMERO PROMEDIO DE DÍAS PERDIDOS POR CADA ACCIDENTE DEL TRABAJO Y DE TRAYECTO SEGÚN ACTIVIDAD ECONÓMICA Y MUTUALIDADES
                    categoryData = {
                        "ACCIDENTES DEL TRABAJO": {
                                "economicActivityStart":'B10',   
                                "economicActivityEnd": 'B27',
                                "totalColumn": 'F',
                                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        }
                    }        
                    data = {}  
                    for category, categoryInfo in categoryData.items():
                        data[category] = {}
                        
                        economicActivityStart = wb[categoryInfo['economicActivityStart']]
                        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
                        totalColumn = categoryInfo['totalColumn']
                        economicActivities = []

                        for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                            economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                            economicActivity = economic_activity_cell.value
                                    
                            achs_value = economic_activity_cell.offset(column=1).value
                            museg_value = economic_activity_cell.offset(column=2).value
                            ist_value = economic_activity_cell.offset(column=3).value
                            totalValue = wb[f"{totalColumn}{row}"].value 
                            
                            additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                                    
                            economicActivities.append({
                                "Economic Activity": economicActivity,
                                "ACHS": achs_value,
                                "MUSEG": museg_value,
                                "IST": ist_value,
                                "TOTAL": totalValue,
                                "Additional Values": additionalValues,
                            })
                        data[category]['Economic Activities'] = economicActivities

                    year_instance, created = Year.objects.get_or_create(year=extracted_year)

                    for category, categoryInfo in data.items():
                        category_instance, created = Category.objects.get_or_create(name=category)
                        category_instances[category] = category_instance

                        for economic_activity_info in categoryInfo['Economic Activities']:
                            activity_name = economic_activity_info['Economic Activity']

                            # Verificar si ya existe una instancia con el mismo nombre
                            economic_activity_instance = economic_activities_instances.get(activity_name)
                            if not economic_activity_instance:
                                # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                                economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                                economic_activities_instances[activity_name] = economic_activity_instance

                            # Crear y guardar la instancia de FallecidosxAct
                            diaxact_instance = DiasxActividad(
                                year=year_instance,
                                category=category_instance,
                                EconomicActivity=economic_activity_instance,
                                achs=economic_activity_info['Additional Values'][0],
                                museg=economic_activity_info['Additional Values'][1],
                                ist=economic_activity_info['Additional Values'][2],
                                total=economic_activity_info['Additional Values'][3],
                            )
                            diaxact_instance.save()

def processingLogicForSheet40(wb, extracted_year):
                print(f"custom logic for sheet: {wb.title}")
                categoryData = {
                    "ACCIDENTES DE TRABAJO": {
                        "economicActivityStart": 'B8',
                        "economicActivityEnd": 'B24',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES DE TRAYECTO":{
                        "economicActivityStart": 'B26',
                        "economicActivityEnd": 'B42',
                        "totalColumn": 'G'
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)":{
                        "economicActivityStart": 'B44',
                        "economicActivityEnd": 'B60',
                        "totalColumn": 'G'
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}

                    economicActivityStart = wb[categoryInfo['economicActivityStart']]
                    economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
                    totalColumn = categoryInfo['totalColumn']

                    economicActivities = []
                    totalValues = []

                    for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
                        economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
                        economicActivity = economicActivityCell.value

                        achsValue = economicActivityCell.offset(column=1).value
                        musegValue = economicActivityCell.offset(column=2).value
                        istValue = economicActivityCell.offset(column=3).value
                        islValue = economicActivityCell.offset(column=4).value
                        totalValue = wb[f"{totalColumn}{row}"].value

                        economicActivities.append({
                            "Economic Activity": economicActivity,
                            "ACHS": achsValue,
                            "MUSEG": musegValue,
                            "IST": istValue,
                            "ISL": islValue
                        })
                        totalValues.append(totalValue)

                    data[category]['Total'] = totalValues
                    data[category]['Economic Activities'] = economicActivities

                year_instance, created = Year.objects.get_or_create(year=extracted_year)

                for category, categoryInfo in data.items():
                    category_instance, created = Category.objects.get_or_create(name=category)
                    category_instances[category] = category_instance

                    for economic_activity_info in categoryInfo['Economic Activities']:
                        activity_name = economic_activity_info['Economic Activity']

                        # Verificar si ya existe una instancia con el mismo nombre
                        economic_activity_instance = economic_activities_instances.get(activity_name)
                        if not economic_activity_instance:
                            # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                            economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                            economic_activities_instances[activity_name] = economic_activity_instance

                        # Crear y guardar la instancia de FallecidosxAct
                        fallecidos_instance = FallecidosxAct(
                            year=year_instance,
                            category=category_instance,
                            EconomicActivity=economic_activity_instance,
                            achs=economic_activity_info['ACHS'],
                            museg=economic_activity_info['MUSEG'],
                            ist=economic_activity_info['IST'],
                            isl=economic_activity_info['ISL'],
                            total=totalValue
                        )
                        fallecidos_instance.save()    

def processingLogicForSheet31(wb, extracted_year):
                print()
                categoryData = {
                    "ACCIDENTES DE TRABAJO": {
                        "mutualInicio": "B8",
                        "mutualFinal": "B11",

                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "mutualInicio": "B13",
                        "mutualFinal": "B16",
                        
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)": {
                        "mutualInicio": "B18",
                        "mutualFinal": "B21",
                    
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    mutualInicioCell = wb[categoryInfo["mutualInicio"]]
                    mutualFinalCell = wb[categoryInfo["mutualFinal"]]
                    
        
                    
                    mutuales = []

                    for row in range(mutualInicioCell.row, mutualFinalCell.row + 1):
                        mutualInicioCell = wb.cell(row=row, column=mutualInicioCell.column)
                        mutualInicio = mutualInicioCell.value
                        
                        anio2017 = mutualInicioCell.offset(column=1).value
                        anio2018 = mutualInicioCell.offset(column=2).value
                        anio2019 = mutualInicioCell.offset(column=3).value
                        anio2020 = mutualInicioCell.offset(column=4).value
                        anio2021 = mutualInicioCell.offset(column=5).value

                        mutuales.append({
                            "Mutualidades": mutualInicio,
                            "2017": anio2017,
                            "2018": anio2018,
                            "2019": anio2019,
                            "2020": anio2020,
                            "2021": anio2021
                        })
                    data[category]['Mutuales'] = mutuales
                        
                for category, categoryInfo in data.items():
                    print(category)
                    for i, mutualInicio in enumerate(categoryInfo['Mutuales']):
                        print(f"mutuales: {mutualInicio['Mutualidades']}")
                        print(f"2018: {mutualInicio['2017']}")
                        print(f"2019: {mutualInicio['2018']}")
                        print(f"2020: {mutualInicio['2019']}")
                        print(f"2021: {mutualInicio['2020']}")
                        print(f"2022: {mutualInicio['2021']}")
                    print()           

def processingLogicForSheet40In2014(wb, extracted_year):
    print("custom logic for the page 34 in 2014")
    categoryData = {
        "ACCIDENTES DE TRABAJO":{
            "economicActivityStart": 'B8',
            "economicActivityEnd": 'B17',
            "totalColumn":'G'
        },
        "ACCIDENTES DE TRAYECTO": {
            "economicActivityStart": 'B19',
            "economicActivityEnd": 'B28',
            "totalColumn": 'G'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "economicActivityStart": 'B30',
            "economicActivityEnd": 'B39',
            "totalColumn": 'G'
        }
    }
    data = {}
    for category, categoryInfo in categoryData.items():
        data[category] = {}

        economicActivityStart = wb[categoryInfo['economicActivityStart']]
        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
        totalColumn = categoryInfo['totalColumn']

        economicActivities = []
        totalValues = []

        for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
            economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
            economicActivity = economicActivityCell.value

            achsValue = economicActivityCell.offset(column=1).value
            cchc = economicActivityCell.offset(column=2).value
            istValue = economicActivityCell.offset(column=3).value
            islValue = economicActivityCell.offset(column=4).value
            totalValue = wb[f"{totalColumn}{row}"].value

            economicActivities.append({
                "Economic Activity": economicActivity,
                "ACHS": achsValue,
                "CCHC": cchc,
                "IST": istValue,
                "ISL": islValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Economic Activities'] = economicActivities

    year_instance, created = Year.objects.get_or_create(year=extracted_year)

    for category, categoryInfo in data.items():
        category_instance, created = Category.objects.get_or_create(name=category)
        category_instances[category] = category_instance

        for economic_activity_info in categoryInfo['Economic Activities']:
            activity_name = economic_activity_info['Economic Activity']

            # Verificar si ya existe una instancia con el mismo nombre
            economic_activity_instance = economic_activities_instances.get(activity_name)
            if not economic_activity_instance:
                # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                economic_activities_instances[activity_name] = economic_activity_instance

            # Crear y guardar la instancia de FallecidosxAct
            fallecidos_instance = FallecidosxAct(
                year=year_instance,
                category=category_instance,
                EconomicActivity=economic_activity_instance,
                achs=economic_activity_info['ACHS'],
                cchc=economic_activity_info['CCHC'],
                ist=economic_activity_info['IST'],
                isl=economic_activity_info['ISL'],
                total=totalValue
            )
            fallecidos_instance.save() 

def processingLogicForSheet40In2015(wb, extracted_year):
    categoryData = {
        "ACCIDENTES DE TRABAJO":{
            "economicActivityStart": 'B8',
            "economicActivityEnd": 'B16',
            "totalColumn": 'G'
        },
        "ACCIDENTES DE TRAYECTO": {
            "economicActivityStart": 'B18',
            "economicActivityEnd": 'B27',
            "totalColumn": 'G'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "economicActivityStart": 'B29',
            "economicActivityEnd": 'B37',
            "totalColumn": 'G'
        }
    }
    data = {}
    for category, categoryInfo in categoryData.items():
        data[category] = {}

        economicActivityStart = wb[categoryInfo['economicActivityStart']]
        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
        totalColumn = categoryInfo['totalColumn']

        economicActivities = []
        totalValues = []

        for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
            economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
            economicActivity = economicActivityCell.value

            achsValue = economicActivityCell.offset(column=1).value
            cchc = economicActivityCell.offset(column=2).value
            istValue = economicActivityCell.offset(column=3).value
            islValue = economicActivityCell.offset(column=4).value
            totalValue = wb[f"{totalColumn}{row}"].value

            economicActivities.append({
                "Economic Activity": economicActivity,
                "ACHS": achsValue,
                "CCHC": cchc,
                "IST": istValue,
                "ISL": islValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Economic Activities'] = economicActivities

    year_instance, created = Year.objects.get_or_create(year=extracted_year)

    for category, categoryInfo in data.items():
        category_instance, created = Category.objects.get_or_create(name=category)
        category_instances[category] = category_instance

        for economic_activity_info in categoryInfo['Economic Activities']:
            activity_name = economic_activity_info['Economic Activity']

            # Verificar si ya existe una instancia con el mismo nombre
            economic_activity_instance = economic_activities_instances.get(activity_name)
            if not economic_activity_instance:
                # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                economic_activities_instances[activity_name] = economic_activity_instance

            # Crear y guardar la instancia de FallecidosxAct
            fallecidos_instance = FallecidosxAct(
                year=year_instance,
                category=category_instance,
                EconomicActivity=economic_activity_instance,
                achs=economic_activity_info['ACHS'],
                cchc=economic_activity_info['CCHC'],
                ist=economic_activity_info['IST'],
                isl=economic_activity_info['ISL'],
                total=totalValue
            )
            fallecidos_instance.save()   

def processigLogicForSheet31In2015(wb, extracted_year):
        categoryData = {
            "ACCIDENTES DE TRABAJO": {
                    "economicActivityStart":'B10',   
                    "economicActivityEnd": 'B27',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES DE TRAYECTO":{
                    "economicActivityStart": 'B29',
                    "economicActivityEnd": 'B46',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES (TRABAJO + TRAYECTO)": {
                "economicActivityStart": 'B48',
                "economicActivityEnd": 'B65',
                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                "totalColumn": 'F',
            }
        }        
        data = {}  
        for category, categoryInfo in categoryData.items():
            data[category] = {}
            
            economicActivityStart = wb[categoryInfo['economicActivityStart']]
            economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
            totalColumn = categoryInfo['totalColumn']
            economicActivities = []

            for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                economicActivity = economic_activity_cell.value
                        
                achs_value = economic_activity_cell.offset(column=1).value
                cchc_value = economic_activity_cell.offset(column=2).value
                ist_value = economic_activity_cell.offset(column=3).value
                totalValue = wb[f"{totalColumn}{row}"].value 
                
                additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                        
                economicActivities.append({
                    "Economic Activity": economicActivity,
                    "ACHS": achs_value,
                    "CCHC": cchc_value,
                    "IST": ist_value,
                    "TOTAL": totalValue,
                    "Additional Values": additionalValues,
                })
            data[category]['Economic Activities'] = economicActivities

        year_instance, created = Year.objects.get_or_create(year=extracted_year)

        for category, categoryInfo in data.items():
            category_instance, created = Category.objects.get_or_create(name=category)
            category_instances[category] = category_instance

            for economic_activity_info in categoryInfo['Economic Activities']:
                activity_name = economic_activity_info['Economic Activity']

                # Verificar si ya existe una instancia con el mismo nombre
                economic_activity_instance = economic_activities_instances.get(activity_name)
                if not economic_activity_instance:
                    # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                    economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                    economic_activities_instances[activity_name] = economic_activity_instance

                # Crear y guardar la instancia de FallecidosxAct
                diaxact_instance = DiasxActividad(
                    year=year_instance,
                    category=category_instance,
                    EconomicActivity=economic_activity_instance,
                    achs=economic_activity_info['Additional Values'][0],
                    cchc=economic_activity_info['Additional Values'][1],
                    ist=economic_activity_info['Additional Values'][2],
                    total=economic_activity_info['Additional Values'][3],
                )
                diaxact_instance.save()

def processLogicForSheet31In2014(wb, extracted_year):
        categoryData = {
            "ACCIDENTES DE TRABAJO": {
                    "economicActivityStart":'B10',   
                    "economicActivityEnd": 'B20',
                    "totalColumn": 'G',
                    "additionalValuesColumns": [ 'H', 'J', 'K', 'L'],
            },
            "ACCIDENTES DE TRAYECTO":{
                    "economicActivityStart": 'B22',
                    "economicActivityEnd": 'B32',
                    "totalColumn": 'G',
                    "additionalValuesColumns": [  'H', 'J', 'K', 'L'],
            },
            "ACCIDENTES (TRABAJO + TRAYECTO)": {
                "economicActivityStart": 'B34',
                "economicActivityEnd": 'B44',
                "additionalValuesColumns": ['H', 'J', 'K', 'L'],
                "totalColumn": 'G',
            }
        }        
        data = {}  
        for category, categoryInfo in categoryData.items():
            data[category] = {}
            
            economicActivityStart = wb[categoryInfo['economicActivityStart']]
            economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
            totalColumn = categoryInfo['totalColumn']
            economicActivities = []

            for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                economicActivity = economic_activity_cell.value
                        
                achs_value = economic_activity_cell.offset(column=1).value
                cchc_value = economic_activity_cell.offset(column=2).value
                ist_value = economic_activity_cell.offset(column=3).value
                totalValue = wb[f"{totalColumn}{row}"].value 
                
                additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                        
                economicActivities.append({
                    "Economic Activity": economicActivity,
                    "ACHS": achs_value,
                    "CCHC": cchc_value,
                    "IST": ist_value,
                    "TOTAL": totalValue,
                    "Additional Values": additionalValues,
                })
            data[category]['Economic Activities'] = economicActivities

        year_instance, created = Year.objects.get_or_create(year=extracted_year)

        for category, categoryInfo in data.items():
            category_instance, created = Category.objects.get_or_create(name=category)
            category_instances[category] = category_instance

            for economic_activity_info in categoryInfo['Economic Activities']:
                activity_name = economic_activity_info['Economic Activity']

                # Verificar si ya existe una instancia con el mismo nombre
                economic_activity_instance = economic_activities_instances.get(activity_name)
                if not economic_activity_instance:
                    # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                    economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                    economic_activities_instances[activity_name] = economic_activity_instance

                # Crear y guardar la instancia de FallecidosxAct
                diaxact_instance = DiasxActividad(
                    year=year_instance,
                    category=category_instance,
                    EconomicActivity=economic_activity_instance,
                    achs=economic_activity_info['Additional Values'][0],
                    cchc=economic_activity_info['Additional Values'][1],
                    ist=economic_activity_info['Additional Values'][2],
                    total=economic_activity_info['Additional Values'][3],
                )
                diaxact_instance.save() 

def processLogicForSheet29(wb, extracted_year):
                print()
                categoryData = {
                    "ACCIDENTES DE TRABAJO": {
                        "mutualInicio": "B8",
                        "mutualFinal": "B11",

                    },
                    "ACCIDENTES DE TRAYECTO": {
                        "mutualInicio": "B13",
                        "mutualFinal": "B16",
                        
                    },
                    "ACCIDENTES (TRABAJO + TRAYECTO)": {
                        "mutualInicio": "B18",
                        "mutualFinal": "B21",
                    
                    }
                }
                data = {}

                for category, categoryInfo in categoryData.items():
                    data[category] = {}
                    
                    mutualInicioCell = wb[categoryInfo["mutualInicio"]]
                    mutualFinalCell = wb[categoryInfo["mutualFinal"]]
                    
        
                    
                    mutuales = []

                    for row in range(mutualInicioCell.row, mutualFinalCell.row + 1):
                        mutualInicioCell = wb.cell(row=row, column=mutualInicioCell.column)
                        mutualInicio = mutualInicioCell.value
                        
                        anio2017 = mutualInicioCell.offset(column=1).value
                        anio2018 = mutualInicioCell.offset(column=2).value
                        anio2019 = mutualInicioCell.offset(column=3).value
                        anio2020 = mutualInicioCell.offset(column=4).value
                        anio2021 = mutualInicioCell.offset(column=5).value

                        mutuales.append({
                            "Mutualidades": mutualInicio,
                            "2018": anio2017,
                            "2019": anio2018,
                            "2020": anio2019,
                            "2021": anio2020,
                            "2022": anio2021
                        })
                    data[category]['Mutuales'] = mutuales
                        
                for category, categoryInfo in data.items():
                    print(category)
                    for i, mutualInicio in enumerate(categoryInfo['Mutuales']):
                        print(f"mutuales: {mutualInicio['Mutualidades']}")
                        print(f"2018: {mutualInicio['2018']}")
                        print(f"2019: {mutualInicio['2019']}")
                        print(f"2020: {mutualInicio['2020']}")
                        print(f"2021: {mutualInicio['2021']}")
                        print(f"2022: {mutualInicio['2022']}")
                    print()    

def processingLogicForSheet41(wb, extracted_year):
    print(f"custom logic for sheet: {wb.title}")
    categoryData={
        "ACCIDENTES DE TRABAJO": {
            "economicActivityStart": 'B8',
            "economicActivityEnd": 'B24',
            "totalColumn": 'E'
        },
        "ACCIDENTES DE TRAYECTO":{
            "economicActivityStart": 'B26',
            "economicActivityEnd": 'B42',
            "totalColumn": 'E'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)":{
            "economicActivityStart": 'B44',
            "economicActivityEnd": 'B60',
            "totalColumn": 'E'
        }
    }
    data = {}

    for category, categoryInfo in categoryData.items():
        data[category] = {}

        economicActivityStart = wb[categoryInfo['economicActivityStart']]
        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
        totalColumn = categoryInfo['totalColumn']

        economicActivities = []
        totalValues = []

        for row in range (economicActivityStart.row, economicActivityEnd.row + 1):
            economicActivityCell = wb.cell(row=row,column=economicActivityStart.column)
            economicActivity = economicActivityCell.value

            menValue = economicActivityCell.offset(column=1).value
            womenValue = economicActivityCell.offset(column=2).value
            totalValue = wb[f"{totalColumn}{row}"].value

            economicActivities.append({
                "Economic Activity": economicActivity,
                "MEN": menValue,
                "WOMEN": womenValue,
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Economic Activities'] = economicActivities

    year_instance, created = Year.objects.get_or_create(year=extracted_year)

    for category, categoryInfo in data.items():
        category_instance, created = Category.objects.get_or_create(name=category)
        category_instances[category] = category_instance

        for economic_activity_info in categoryInfo['Economic Activities']:
            activity_name = economic_activity_info['Economic Activity']

            # Verificar si ya existe una instancia con el mismo nombre
            economic_activity_instance = economic_activities_instances.get(activity_name)
            if not economic_activity_instance:
                # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                economic_activities_instances[activity_name] = economic_activity_instance

            # Crear y guardar la instancia de FallecidosxAct
            fallecidosxSexo_instance = FallecidosxSexo(
                year=year_instance,
                category=category_instance,
                EconomicActivity=economic_activity_instance,
                men=economic_activity_info['MEN'],
                women=economic_activity_info['WOMEN'],
                total=totalValue
            )
            fallecidosxSexo_instance.save()  
          
def processLogicForSheet30(wb, extracted_year):
        categoryData = {
            "ACCIDENTES DE TRABAJO": {
                    "economicActivityStart":'B10',   
                    "economicActivityEnd": 'B27',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES DE TRAYECTO":{
                    "economicActivityStart": 'B29',
                    "economicActivityEnd": 'B46',
                    "totalColumn": 'F',
                    "additionalValuesColumns": ['G', 'H', 'I', 'J'],
            },
            "ACCIDENTES (TRABAJO + TRAYECTO)": {
                "economicActivityStart": 'B48',
                "economicActivityEnd": 'B65',
                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                "totalColumn": 'F',
            }
        }        
        data = {}  
        for category, categoryInfo in categoryData.items():
            data[category] = {}
            
            economicActivityStart = wb[categoryInfo['economicActivityStart']]
            economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
            totalColumn = categoryInfo['totalColumn']
            economicActivities = []

            for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                economicActivity = economic_activity_cell.value
                        
                achs_value = economic_activity_cell.offset(column=1).value
                museg_value = economic_activity_cell.offset(column=2).value
                ist_value = economic_activity_cell.offset(column=3).value
                totalValue = wb[f"{totalColumn}{row}"].value 
                
                additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                        
                economicActivities.append({
                    "Economic Activity": economicActivity,
                    "ACHS": achs_value,
                    "MUSEG": museg_value,
                    "IST": ist_value,
                    "TOTAL": totalValue,
                    "Additional Values": additionalValues,
                })
            data[category]['Economic Activities'] = economicActivities

        year_instance, created = Year.objects.get_or_create(year=extracted_year)

        for category, categoryInfo in data.items():
            category_instance, created = Category.objects.get_or_create(name=category)
            category_instances[category] = category_instance

            for economic_activity_info in categoryInfo['Economic Activities']:
                activity_name = economic_activity_info['Economic Activity']

                # Verificar si ya existe una instancia con el mismo nombre
                economic_activity_instance = economic_activities_instances.get(activity_name)
                if not economic_activity_instance:
                    # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                    economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                    economic_activities_instances[activity_name] = economic_activity_instance

                # Crear y guardar la instancia de FallecidosxAct
                tasasxact_instance = TasaxAct(
                    year=year_instance,
                    category=category_instance,
                    EconomicActivity=economic_activity_instance,
                    achs=economic_activity_info['Additional Values'][0],
                    museg=economic_activity_info['Additional Values'][1],
                    ist=economic_activity_info['Additional Values'][2],
                    total=economic_activity_info['Additional Values'][3],
                )
                tasasxact_instance.save()

def processLogicForSheet30In2014(wb, extracted_year):
                    categoryData = {
                        "ACCIDENTES DE TRABAJO": {
                                "economicActivityStart":'B10',   
                                "economicActivityEnd": 'B27',
                                "totalColumn": 'F',
                                "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        },
                        "ACCIDENTES DE TRAYECTO": {
                            "economicActivityStart":'B29',   
                            "economicActivityEnd": 'B46',
                            "totalColumn": 'F',
                            "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        },
                        "TASA ACCIDENTABILIDAD": {
                            "economicActivityStart":'B48',   
                            "economicActivityEnd": 'B65',
                            "totalColumn": 'F',
                            "additionalValuesColumns": ['G', 'H', 'I', 'J'],
                        }
                    }        
                    data = {}  
                    for category, categoryInfo in categoryData.items():
                        data[category] = {}
                        
                        economicActivityStart = wb[categoryInfo['economicActivityStart']]
                        economicActivityEnd = wb[categoryInfo['economicActivityEnd']]
                        totalColumn = categoryInfo['totalColumn']
                        economicActivities = []

                        for row in range(economicActivityStart.row, economicActivityEnd.row + 1):
                            economic_activity_cell = wb.cell(row=row, column=economicActivityStart.column)
                            economicActivity = economic_activity_cell.value
                                    
                            achs_value = economic_activity_cell.offset(column=1).value
                            cchc_value = economic_activity_cell.offset(column=2).value
                            ist_value = economic_activity_cell.offset(column=3).value
                            totalValue = wb[f"{totalColumn}{row}"].value 
                            
                            additionalValues = [wb[f"{col_letter}{row}"].value for col_letter in categoryInfo.get("additionalValuesColumns", [])]
                                    
                            economicActivities.append({
                                "Economic Activity": economicActivity,
                                "ACHS": achs_value,
                                "CCHC": cchc_value,
                                "IST": ist_value,
                                "TOTAL": totalValue,
                                "Additional Values": additionalValues,
                            })
                        data[category]['Economic Activities'] = economicActivities

                    year_instance, created = Year.objects.get_or_create(year=extracted_year)

                    for category, categoryInfo in data.items():
                        category_instance, created = Category.objects.get_or_create(name=category)
                        category_instances[category] = category_instance

                        for economic_activity_info in categoryInfo['Economic Activities']:
                            activity_name = economic_activity_info['Economic Activity']

                            # Verificar si ya existe una instancia con el mismo nombre
                            economic_activity_instance = economic_activities_instances.get(activity_name)
                            if not economic_activity_instance:
                                # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                                economic_activity_instance, created = EconomicActivity.objects.get_or_create(activity_name=activity_name)
                                economic_activities_instances[activity_name] = economic_activity_instance

                            # Crear y guardar la instancia de FallecidosxAct
                            tasasxact_instance = TasaxAct(
                                year=year_instance,
                                category=category_instance,
                                EconomicActivity=economic_activity_instance,
                                achs=economic_activity_info['Additional Values'][0],
                                cchc=economic_activity_info['Additional Values'][1],
                                ist=economic_activity_info['Additional Values'][2],
                                total=economic_activity_info['Additional Values'][3],
                            )
                            tasasxact_instance.save()

region_instances = {}

def processingLogicForSheet25(wb, extracted_year):
    categoryData = {
        "ACCIDENTES DE TRABAJO":{
            "regionStart": 'B9',
            "regionEnd": 'B25',
            "totalColumn": 'F'
        },
        "ACCIDENTES DE TRAYECTO":{
            "regionStart": 'B27',
            "regionEnd": 'B43',
            "totalColumn": 'F'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "regionStart": 'B45',
            "regionEnd": 'B61',
            "totalColumn": 'F'
        },
        "ENFERMEDADES PROFESIONALES": {
            "regionStart": 'B63',
            "regionEnd": 'B79',
            "totalColumn": 'F'
        }
    }
    data = {}

    for category, categoryInfo in categoryData.items():
        data[category] = {}

        regionStart = wb[categoryInfo['regionStart']]
        regionEnd = wb[categoryInfo['regionEnd']]
        totalColumn = categoryInfo['totalColumn']

        regionActivities = []
        totalValues = []

        for row in range (regionStart.row, regionEnd.row + 1):
            regionStartCell = wb.cell(row=row,column=regionStart.column)
            regionActivity = regionStartCell.value

            achsValue = regionStartCell.offset(column=1).value
            musegValue = regionStartCell.offset(column=2).value
            istValue = regionStartCell.offset(column=3).value
            totalValue = wb[f"{totalColumn}{row}"].value

            regionActivities.append({
                "Activity region": regionActivity,
                "ACHS": achsValue,
                "MUSEG": musegValue,
                "IST": istValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Region Activities'] = regionActivities

    year_instance, created = Year.objects.get_or_create(year=extracted_year)

    for category, categoryInfo in data.items():
        category_instance, created = Category.objects.get_or_create(name=category)
        category_instances[category] = category_instance

        for region_info in categoryInfo['Region Activities']:
            region = region_info['Activity region']

            # Verificar si ya existe una instancia con el mismo nombre
            region_instance = region_instances.get(region)
            if not region_instance:
                # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                region_instance, created = Region.objects.get_or_create(region=region)
                region_instances[region] = region_instance

            # Crear y guardar la instancia de AccidentesxRegion
            accidentexRegion_instance = AccidentesxRegion(
                year=year_instance,
                category=category_instance,
                region=region_instance,
                achs=region_info['ACHS'],
                museg=region_info['MUSEG'],
                ist=region_info['IST'],
                total=totalValue  # Usar el valor correcto desde la lista totalValues
            )
            accidentexRegion_instance.save()   
        
def processingLogicForSheet25v2(wb, extracted_year):
    categoryData = {
        "ACCIDENTES DE TRABAJO":{
            "regionStart": 'B9',
            "regionEnd": 'B25',
            "totalColumn": 'F'
        },
        "ACCIDENTES DE TRAYECTO":{
            "regionStart": 'B27',
            "regionEnd": 'B43',
            "totalColumn": 'F'
        },
        "ACCIDENTES (TRABAJO + TRAYECTO)": {
            "regionStart": 'B45',
            "regionEnd": 'B61',
            "totalColumn": 'F'
        },
        "ENFERMEDADES PROFESIONALES": {
            "regionStart": 'B63',
            "regionEnd": 'B79',
            "totalColumn": 'F'
        }
    }
    data = {}

    for category, categoryInfo in categoryData.items():
        data[category] = {}

        regionStart = wb[categoryInfo['regionStart']]
        regionEnd = wb[categoryInfo['regionEnd']]
        totalColumn = categoryInfo['totalColumn']

        regionActivities = []
        totalValues = []

        for row in range (regionStart.row, regionEnd.row + 1):
            regionStartCell = wb.cell(row=row,column=regionStart.column)
            regionActivity = regionStartCell.value

            achsValue = regionStartCell.offset(column=1).value
            musegValue = regionStartCell.offset(column=2).value
            istValue = regionStartCell.offset(column=3).value
            totalValue = wb[f"{totalColumn}{row}"].value

            regionActivities.append({
                "Activity region": regionActivity,
                "ACHS": achsValue,
                "MUSEG": musegValue,
                "IST": istValue
            })
            totalValues.append(totalValue)

        data[category]['Total'] = totalValues
        data[category]['Region Activities'] = regionActivities

    year_instance, created = Year.objects.get_or_create(year=extracted_year)

    for category, categoryInfo in data.items():
        category_instance, created = Category.objects.get_or_create(name=category)
        category_instances[category] = category_instance

        for region_info in categoryInfo['Region Activities']:
            region = region_info['Activity region']

            # Verificar si ya existe una instancia con el mismo nombre
            region_instance = region_instances.get(region)
            if not region_instance:
                # Si no existe, se crea una nueva instancia y se guarda en el diccionario
                region_instance, created = Region.objects.get_or_create(region=region)
                region_instances[region] = region_instance

            # Crear y guardar la instancia de AccidentesxRegion
            accidentexRegion_instance = AccidentesxRegion(
                year=year_instance,
                category=category_instance,
                region=region_instance,
                achs=region_info['ACHS'],
                museg=region_info['MUSEG'],
                ist=region_info['IST'],
                total=totalValue  # Usar el valor correcto desde la lista totalValues
            )
            accidentexRegion_instance.save()


    
def getFileYear(fileName):
    #decodedFileName = fileName.encode('utf-8').decode('utf-8')
    cleanedFileName = unidecode(fileName)
    match = re.search(r'\d{4}', cleanedFileName)
    if match:
        year = int(match.group())
        print(f"Extracted year: {year}")
        return year
    return None
 
def is_2022_file(fileName):
    
    #cleanedFileName = unidecode.unidecode(fileName)
    cleanedFileName = unidecode(fileName)
    pattern = re.compile(r'2022', re.IGNORECASE)
    result =  re.search(pattern, cleanedFileName) is not None
    print(f"Original File Name: {fileName}")
    print(f"Cleaned File Name: {cleanedFileName}")
    print(f"Checking if {fileName} is a 2022 file: {result}")
    return result

if __name__ == "__main__":
    webPageUrl = "https://www.suseso.cl/608/w3-propertyvalue-10364.html"# Replace with your URL
    downloadDir = os.path.join(os.getcwd(), "downloadedFiles")
    
    
    downloadedFiles = downloadExcelFiles(webPageUrl, downloadDir)
    downloadedFiles.sort(key=lambda x: getFileYear(x[0]) or 0, reverse=True)

    
    lastProcessedFileYear = None
    

    
    for title, filePath in downloadedFiles:
        if is_2022_file(title):
            print(f"Identified as 2022. Processing with type1")
            extractedDataFromExcel_Type1(filePath)
            lastProcessedFileYear = 2022
        else:
            year = getFileYear(title)
            if lastProcessedFileYear is None or year == lastProcessedFileYear - 1:
                extractedDataFromExcel_Type2(filePath, year)
                lastProcessedFileYear = year
            else:
                print(f"Skipping file: {title} (Unexpected year)")